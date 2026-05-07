import urllib.request, urllib.parse, json, io
import openpyxl

data = urllib.parse.urlencode({'matricule': '9999', 'password': 'ChangeMe123!@#'}).encode()
req = urllib.request.Request('http://localhost:8000/auth/login', data=data, method='POST')
token = json.loads(urllib.request.urlopen(req).read())['access_token']

# Check global export sheets
body = urllib.request.urlopen(
    urllib.request.Request('http://localhost:8000/api/analytics/export-excel',
    headers={'Authorization': f'Bearer {token}'})
).read()
wb = openpyxl.load_workbook(io.BytesIO(body))
print('Global export sheets:', wb.sheetnames)

# Check dashboard export sheets
body2 = urllib.request.urlopen(
    urllib.request.Request('http://localhost:8000/api/analytics/export-dashboard/9999',
    headers={'Authorization': f'Bearer {token}'})
).read()
wb2 = openpyxl.load_workbook(io.BytesIO(body2))
print('Dashboard export sheets:', wb2.sheetnames)
