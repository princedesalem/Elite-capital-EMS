# -*- coding: utf-8 -*-
"""
Suivi EMS — génère le fichier Excel de suivi de fonctionnalités.
Toutes les valeurs dynamiques (% par ligne, TOTAL, Tableau de Bord) sont
des FORMULES Excel : changer l'état dans la liste déroulante met
automatiquement à jour la couleur, le %, le total et le tableau de bord.
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Charte EMS (frontend/src/index.css) ───────────────────────────────────
BLEU       = "02162E"
BLEU_CLAIR = "17283D"
ROUGE      = "D0202B"
GRIS       = "606060"
BLANC      = "FFFFFF"
FOND_APP   = "F4F5F7"

# (bg, fg) par état
ETAT_TERMINE  = ("E8F5E9", BLEU)
ETAT_EN_COURS = (BLEU_CLAIR, BLANC)
ETAT_NON_DEB  = ("F8D7DA", ROUGE)
STATUS_MAP = {
    "Terminé":    ETAT_TERMINE,
    "En cours":   ETAT_EN_COURS,
    "Non débuté": ETAT_NON_DEB,
}
STATUS_PCT = {"Terminé": 1.0, "En cours": 0.5, "Non débuté": 0.0}
ROW_FILLS  = [BLANC, FOND_APP]

# ── Colonnes ─────────────────────────────────────────────────────────────
COLS = [
    ("Rubrique",         28),
    ("Action",           55),
    ("Date d'insertion", 16),
    ("État",             16),
    ("Délais",           10),
    ("Date de fin",      14),
    ("% Avancement",     14),
]
N_COLS   = len(COLS)
PCT_COL  = N_COLS                        # 7
PCT_LTR  = get_column_letter(PCT_COL)    # "G"

# ── Données ──────────────────────────────────────────────────────────────
TODAY = "27/04/2026"

ROWS = [
    ("Organigramme",
     "Le responsable doit toujours être plus haut que l'employé dans le visuel de l'organigramme",
     TODAY, "En cours", "3 j", ""),
    ("Missions",
     "Bloquer la clôture d'une mission si une preuve de frais n'est pas téléversée",
     TODAY, "Terminé", "1 j", "28/04/2026"),
    ("Missions",
     "Notifications de frais : libellé → \"Une nouvelle demande...\" ; bouton → \"Voir la demande\"",
     TODAY, "Non débuté", "1 j", ""),
    ("Missions",
     "Supprimer \"Nouvelle mission multi-destinations\" → utiliser \"Nouvelle mission\"",
     TODAY, "Terminé", "0.5 j", "28/04/2026"),
    ("Employés",
     "Rendre le matricule alphanumérique (lettres + chiffres)",
     TODAY, "Terminé", "1 j", "28/04/2026"),
    ("Employés",
     "Enregistrer le salaire à la création ; chaque employé voit uniquement le sien ; RH voit tout",
     TODAY, "En cours", "2 j", ""),
    ("Employés",
     "Ajouter le filtrage des employés par sexe",
     TODAY, "Terminé", "0.5 j", "28/04/2026"),
    ("Opérations",
     "Les opérations doivent apparaître simultanément chez les deux DG",
     TODAY, "Terminé", "1 j", "28/04/2026"),
    ("Notifications",
     "Relances et notifications envoyées hors application (email + mobile push)",
     TODAY, "En cours", "5 j", ""),
    ("Notifications",
     "Supprimer la mention \"en tant que [rôle]\" dans toutes les notifications",
     TODAY, "Non débuté", "0.5 j", ""),
    ("Notifications",
     "Toast de notification : utiliser le bleu de la charte graphique EMS",
     TODAY, "Terminé", "0.5 j", "28/04/2026"),
    ("Biométrie / Intégration",
     "Prévoir la connexion avec une application de biométrie pour gérer les retards",
     TODAY, "Non débuté", "10 j", ""),
    ("Biométrie / Intégration",
     "Intégration Sage Paie (automatisation des retenues salariales)",
     TODAY, "Non débuté", "15 j", ""),
    ("Congés",
     "Règle : demande créée > 3 sem avant début → modification bloquée à 2 sem avant début",
     TODAY, "Terminé", "1 j", "28/04/2026"),
    ("Congés / Permissions",
     "Maternité simple : 14 semaines (était 16) ; Maternité pathologique : 20 semaines (était 18)",
     TODAY, "Terminé", "0.5 j", "28/04/2026"),
    ("UI / UX",
     "Ajouter un scroll horizontal sur chaque tableau de l'application",
     TODAY, "Terminé", "1 j", "28/04/2026"),
    ("UI / UX",
     "Ajouter \"et autres\" à côté du dernier pays dans l'onglet Organisation",
     TODAY, "Terminé", "0.5 j", "28/04/2026"),
    ("Sécurité",
     "Politique de gestion des failles zero-day (procédure, veille CVE, patch management)",
     TODAY, "En cours", "5 j", ""),
    ("Performance & Évaluation",
     "Notation : respect délais validation (valideur max 3h, sinon mauvaise note)",
     TODAY, "Non débuté", "7 j", ""),
    ("Performance & Évaluation",
     "Notation : comportement sur l'application",
     TODAY, "Non débuté", "5 j", ""),
    ("Performance & Évaluation",
     "Notation : taux de participation aux événements (liste participation + présence le jour J)",
     TODAY, "Non débuté", "7 j", ""),
    ("Performance & Évaluation",
     "Intégrer les notations dans Rubrique Parcours, Module Performance, 360°",
     TODAY, "Non débuté", "10 j", ""),
    ("Performance & Évaluation",
     "Évaluation esprit d'équipe",
     TODAY, "Non débuté", "5 j", ""),
    ("Gestion Disciplinaire",
     "Module : blâme, avertissement, sanctions, demande d'explication, conseil de discipline",
     TODAY, "Non débuté", "10 j", ""),
    ("Gestion Disciplinaire",
     "Ajouter un module Demande d'Explication (DE) dédié",
     TODAY, "Non débuté", "5 j", ""),
    ("Reporting & Data",
     "Export Excel complet incluant les données sources des graphiques",
     TODAY, "Non débuté", "3 j", ""),
    ("Reporting & Data",
     "Amélioration PDF : template avec logo en haut à gauche sur tous les exports PDF",
     TODAY, "Non débuté", "3 j", ""),
    ("Reporting & Data",
     "PDF : signature automatique des validateurs lors de chaque validation",
     TODAY, "Non débuté", "4 j", ""),
    ("Reporting & Data",
     "PDF : ajouter une table de signatures",
     TODAY, "Non débuté", "2 j", ""),
    ("IA / Aide à la Décision",
     "Résumé automatique du dashboard par IA",
     TODAY, "Non débuté", "15 j", ""),
    ("IA / Aide à la Décision",
     "Recommandations contextuelles pour managers (solde congés critique, actions suggérées)",
     TODAY, "Non débuté", "10 j", ""),
    ("IA / Aide à la Décision",
     "Rapport narratif PDF en langage naturel pour comités de direction",
     TODAY, "Non débuté", "12 j", ""),
    ("NLP & Recherche",
     "Recherche employés en langage naturel → génération SQL automatique",
     TODAY, "Non débuté", "15 j", ""),
    ("NLP & Recherche",
     "Chatbot RH interne (solde congés, statut missions, etc.)",
     TODAY, "Non débuté", "20 j", ""),
    ("NLP & Recherche",
     "Extraction d'informations depuis documents/contrats pour pré-remplir les formulaires",
     TODAY, "Non débuté", "15 j", ""),
    ("Analyse Performance",
     "Scoring composite par département (présence + productivité + missions + formations)",
     TODAY, "Non débuté", "12 j", ""),
    ("Analyse Performance",
     "Détection de biais dans les validations congés/permissions (hommes/femmes, entités, grades)",
     TODAY, "Non débuté", "10 j", ""),
    ("Tests",
     "Ajouter des tests automatisés pour toutes les fonctionnalités ci-dessus",
     TODAY, "Non débuté", "20 j", ""),
]
N_ROWS = len(ROWS)

# ═══════════════════════════════════════════════════════════════════════════
# ONGLET 1 — Suivi EMS
# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Suivi EMS"

thin   = Side(style="thin",   color="E5E7EB")
medium = Side(style="medium", color=ROUGE)
b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
b_med  = Border(left=medium, right=medium, top=medium, bottom=medium)

# Ligne 1 — titre
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
t = ws.cell(row=1, column=1, value="SUIVI DES FONCTIONNALITÉS EMS — Roadmap 2026")
t.fill      = PatternFill("solid", fgColor=BLEU)
t.font      = Font(bold=True, color=BLANC, name="Calibri", size=14)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

# Ligne 2 — en-têtes
for ci, (name, width) in enumerate(COLS, start=1):
    c = ws.cell(row=2, column=ci, value=name)
    c.fill      = PatternFill("solid", fgColor=BLEU_CLAIR)
    c.font      = Font(bold=True, color=BLANC, name="Calibri", size=11)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = b_thin
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[2].height = 30
ws.freeze_panes = "A3"

# Lignes 3..N+2 — données
DATA_FIRST = 3
DATA_LAST  = DATA_FIRST + N_ROWS - 1

prev_rub, fill_idx = None, 0
for i, (rub, action, date_ins, etat, delais, date_fin) in enumerate(ROWS):
    row = DATA_FIRST + i
    if rub != prev_rub:
        fill_idx = 1 - fill_idx
        prev_rub = rub
    row_bg = ROW_FILLS[fill_idx]
    etat_bg, etat_fg = STATUS_MAP.get(etat, (row_bg, GRIS))
    values = [rub, action, date_ins, etat, delais, date_fin, None]

    for ci, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.border    = b_thin
        c.alignment = Alignment(vertical="top", wrap_text=True)

        if ci == 1:
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.font = Font(name="Calibri", size=10, bold=True, color=BLEU)
        elif ci == 4:  # État
            c.fill      = PatternFill("solid", fgColor=etat_bg)
            c.font      = Font(name="Calibri", size=10, bold=True, color=etat_fg)
            c.alignment = Alignment(horizontal="center", vertical="center")
        elif ci == 7:  # % — formule
            c.value         = f'=IF(D{row}="Terminé",1,IF(D{row}="En cours",0.5,0))'
            c.fill          = PatternFill("solid", fgColor=etat_bg)
            c.font          = Font(name="Calibri", size=10, bold=True, color=etat_fg)
            c.number_format = "0%"
            c.alignment     = Alignment(horizontal="center", vertical="center")
        elif ci in (3, 5, 6):
            c.fill      = PatternFill("solid", fgColor=row_bg)
            c.font      = Font(name="Calibri", size=10, color=GRIS)
            c.alignment = Alignment(horizontal="center", vertical="top")
        else:
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.font = Font(name="Calibri", size=10, color=GRIS)
    ws.row_dimensions[row].height = 36

# Ligne TOTAL — formules COUNTIF / AVERAGE
TOTAL_ROW  = DATA_LAST + 1
etat_range = f"D{DATA_FIRST}:D{DATA_LAST}"
pct_range  = f"{PCT_LTR}{DATA_FIRST}:{PCT_LTR}{DATA_LAST}"

ws.merge_cells(start_row=TOTAL_ROW, start_column=1, end_row=TOTAL_ROW, end_column=3)
lbl = ws.cell(row=TOTAL_ROW, column=1)
lbl.value = (
    f'="TOTAL  —  {N_ROWS} tâches  |  "'
    f'&COUNTIF({etat_range},"Terminé")&" Terminé  |  "'
    f'&COUNTIF({etat_range},"En cours")&" En cours  |  "'
    f'&COUNTIF({etat_range},"Non débuté")&" Non débuté"'
)
lbl.fill      = PatternFill("solid", fgColor=BLEU)
lbl.font      = Font(name="Calibri", size=11, bold=True, color=BLANC)
lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
lbl.border    = b_med

for ci in range(4, PCT_COL):
    c = ws.cell(row=TOTAL_ROW, column=ci)
    c.fill   = PatternFill("solid", fgColor=BLEU)
    c.border = b_med

# % global = AVERAGE des cellules G (chaque G ∈ {0, 0.5, 1})
pct_total = ws.cell(row=TOTAL_ROW, column=PCT_COL)
pct_total.value         = f"=AVERAGE({pct_range})"
pct_total.fill          = PatternFill("solid", fgColor=BLEU)
pct_total.font          = Font(name="Calibri", size=14, bold=True, color=BLANC)
pct_total.number_format = "0.0%"
pct_total.alignment     = Alignment(horizontal="center", vertical="center")
pct_total.border        = b_med
ws.row_dimensions[TOTAL_ROW].height = 38

# Liste déroulante colonne D
dv = DataValidation(
    type="list",
    formula1='"Terminé,En cours,Non débuté"',
    allow_blank=False,
    showDropDown=False,   # False ⇒ flèche VISIBLE
    showErrorMessage=True,
    errorTitle="Valeur invalide",
    error="Choisissez : Terminé  |  En cours  |  Non débuté",
)
dv.sqref = f"D{DATA_FIRST}:D{DATA_LAST}"
ws.add_data_validation(dv)

# Mise en forme conditionnelle — colonnes D ET G
for col_ltr in ("D", PCT_LTR):
    rng = f"{col_ltr}{DATA_FIRST}:{col_ltr}{DATA_LAST}"
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'$D{DATA_FIRST}="Terminé"'],
        fill=PatternFill("solid", fgColor=ETAT_TERMINE[0]),
        font=Font(name="Calibri", size=10, bold=True, color=ETAT_TERMINE[1]),
        stopIfTrue=True,
    ))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'$D{DATA_FIRST}="En cours"'],
        fill=PatternFill("solid", fgColor=ETAT_EN_COURS[0]),
        font=Font(name="Calibri", size=10, bold=True, color=ETAT_EN_COURS[1]),
        stopIfTrue=True,
    ))
    ws.conditional_formatting.add(rng, FormulaRule(
        formula=[f'$D{DATA_FIRST}="Non débuté"'],
        fill=PatternFill("solid", fgColor=ETAT_NON_DEB[0]),
        font=Font(name="Calibri", size=10, bold=True, color=ETAT_NON_DEB[1]),
        stopIfTrue=True,
    ))

ws.auto_filter.ref = f"A2:{get_column_letter(N_COLS)}2"

# ═══════════════════════════════════════════════════════════════════════════
# ONGLET 2 — Tableau de Bord (100 % formules)
# ═══════════════════════════════════════════════════════════════════════════
wb2 = wb.create_sheet(title="Tableau de Bord")

def style(c, bg, fg, size=11, bold=True, h="center", border=None):
    c.fill      = PatternFill("solid", fgColor=bg)
    c.font      = Font(name="Calibri", size=size, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=h, vertical="center")
    if border:
        c.border = border

med_b  = Border(left=medium, right=medium, top=medium, bottom=medium)
thin_b = Border(left=thin,   right=thin,   top=thin,   bottom=thin)

# Titre
wb2.merge_cells("A1:F1")
style(wb2["A1"], BLEU, BLANC, size=16)
wb2["A1"].value = "TABLEAU DE BORD — Avancement du Projet EMS"
wb2.row_dimensions[1].height = 45
wb2.row_dimensions[2].height = 16

# Référence vers la cellule TOTAL %
TOTAL_REF = f"'Suivi EMS'!{PCT_LTR}{TOTAL_ROW}"
ETAT_REF  = f"'Suivi EMS'!{etat_range}"

# KPI principal — % global
wb2.merge_cells("B3:F3")
wb2.merge_cells("B4:F4")
lbl3 = wb2.cell(row=3, column=2, value="AVANCEMENT GLOBAL DU PROJET")
style(lbl3, BLEU, BLANC, size=13, border=med_b)
val4 = wb2.cell(row=4, column=2)
val4.value         = f"={TOTAL_REF}"
val4.number_format = "0.0%"
style(val4, BLEU, BLANC, size=32, border=med_b)
wb2.row_dimensions[3].height = 28
wb2.row_dimensions[4].height = 60
wb2.row_dimensions[5].height = 16

# Détail par état
wb2.row_dimensions[6].height = 24
wb2.row_dimensions[7].height = 42

# Terminé
wb2.merge_cells("B6:C6"); wb2.merge_cells("B7:C7")
lbl_t = wb2.cell(row=6, column=2,
    value=f'="Terminé  —  "&COUNTIF({ETAT_REF},"Terminé")&" / {N_ROWS}"')
val_t = wb2.cell(row=7, column=2)
val_t.value         = f'=COUNTIF({ETAT_REF},"Terminé")/{N_ROWS}'
val_t.number_format = "0.0%"
style(lbl_t, ETAT_TERMINE[0], ETAT_TERMINE[1], border=thin_b)
style(val_t, ETAT_TERMINE[0], ETAT_TERMINE[1], size=22, border=thin_b)

# En cours
wb2.merge_cells("D6:E6"); wb2.merge_cells("D7:E7")
lbl_e = wb2.cell(row=6, column=4,
    value=f'="En cours  —  "&COUNTIF({ETAT_REF},"En cours")&" / {N_ROWS}"')
val_e = wb2.cell(row=7, column=4)
val_e.value         = f'=COUNTIF({ETAT_REF},"En cours")/{N_ROWS}'
val_e.number_format = "0.0%"
style(lbl_e, ETAT_EN_COURS[0], ETAT_EN_COURS[1], border=thin_b)
style(val_e, ETAT_EN_COURS[0], ETAT_EN_COURS[1], size=22, border=thin_b)

# Non débuté
lbl_n = wb2.cell(row=6, column=6,
    value=f'="Non débuté  —  "&COUNTIF({ETAT_REF},"Non débuté")&" / {N_ROWS}"')
val_n = wb2.cell(row=7, column=6)
val_n.value         = f'=COUNTIF({ETAT_REF},"Non débuté")/{N_ROWS}'
val_n.number_format = "0.0%"
style(lbl_n, ETAT_NON_DEB[0], ETAT_NON_DEB[1], border=thin_b)
style(val_n, ETAT_NON_DEB[0], ETAT_NON_DEB[1], size=22, border=thin_b)

wb2.row_dimensions[8].height = 16

# Bandeau % global en grand
wb2.merge_cells("B9:F9")
bar = wb2.cell(row=9, column=2)
bar.value         = f"={TOTAL_REF}"
bar.number_format = '0.0%" — avancement global du projet"'
style(bar, BLEU_CLAIR, BLANC, size=18, border=thin_b)
wb2.row_dimensions[9].height = 40

# Note
wb2.merge_cells("B11:F11")
note = wb2.cell(row=11, column=2,
    value=f"Dernière mise à jour : {TODAY}  |  Total tâches : {N_ROWS}  |  "
          f"Modifiez l'état dans l'onglet « Suivi EMS » colonne D")
note.font      = Font(name="Calibri", size=9, italic=True, color=GRIS)
note.alignment = Alignment(horizontal="center")

wb2.column_dimensions["A"].width = 3
for col in "BCDEF":
    wb2.column_dimensions[col].width = 22

# Suivi EMS en premier
wb.move_sheet("Tableau de Bord", offset=1)

# ── Save ─────────────────────────────────────────────────────────────────
out_path = (
    r"c:\Users\cedri\OneDrive - ELITE CAPITAL Group S.A"
    r"\Documents\EMS\extranet\SUIVI_FONCTIONNALITES_EMS_2026.xlsx"
)
try:
    wb.save(out_path)
except PermissionError:
    out_path = out_path.replace(".xlsx", "_NEW.xlsx")
    wb.save(out_path)
    print("⚠  Fichier d'origine ouvert → enregistré sous une variante :")

n_t = sum(1 for r in ROWS if r[3] == "Terminé")
n_e = sum(1 for r in ROWS if r[3] == "En cours")
n_n = sum(1 for r in ROWS if r[3] == "Non débuté")
pct = round((n_t * 100 + n_e * 50) / (N_ROWS * 100) * 100, 1)
print(f"Saved   → {out_path}")
print(f"Initial → {pct}%  ({n_t} terminé / {n_e} en cours / {n_n} non débuté) sur {N_ROWS} tâches")
print("► Toutes les valeurs (TOTAL, %, Tableau de Bord) sont des FORMULES Excel :")
print("  changer l'état dans la liste déroulante met tout à jour automatiquement.")
