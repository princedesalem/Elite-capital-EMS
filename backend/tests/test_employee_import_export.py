from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from app.routers.employees import AccessImportError
from app.utils.employee_excel import COLUMNS, workbook_bytes


def test_import_employees_csv(client, seed_reference_data, auth_headers):
    csv_content = (
        "matricule,nom,prenom,email,sexe,date_embauche,entite,direction,departement,role,fonction,contact_urgence\n"
        "8888,Doe,John,8888@example.com,M,2026-02-01,ELCAM,Direction Generale,Operations,EMPLOYE,Analyste,+237699000000\n"
    )
    files = {
        'file': ('employees.csv', csv_content.encode('utf-8'), 'text/csv')
    }

    response = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200

    payload = response.json()
    assert payload['total_rows'] == 1
    assert payload['imported_rows'] == 1
    assert payload['failed_rows'] == 0

    check = client.get('/employees/8888')
    assert check.status_code == 200
    assert check.json()['nom'] == 'Doe'
    assert check.json()['contact_urgence'] == '+237699000000'


def test_import_employees_xlsx(client, seed_reference_data, auth_headers):
    df = pd.DataFrame([
        {
            'matricule': 8889,
            'nom': 'Smith',
            'prenom': 'Jane',
            'email': '8889@example.com',
            'sexe': 'F',
            'date_embauche': date(2026, 2, 2),
            'entite': 'ELCAM',
            'direction': 'Direction Generale',
            'departement': 'Operations',
            'role': 'EMPLOYE',
            'fonction': 'Analyste',
            'contact_urgence': '+237688000000',
        }
    ])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)

    files = {
        'file': ('employees.xlsx', output.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }

    response = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200
    assert response.json()['imported_rows'] == 1

    check = client.get('/employees/8889')
    assert check.status_code == 200
    assert check.json()['contact_urgence'] == '+237688000000'


def test_import_employees_styled_xlsx_with_new_fields(client, seed_reference_data, auth_headers):
    stream = workbook_bytes(rows=[{
        'matricule': '8893',
        'nom': 'Styled',
        'prenom': 'Import',
        'email': '8893@example.com',
        'telephone': '+237611223344',
        'sexe': 'F',
        'date_naissance': '1996-01-15',
        'date_embauche': '2026-02-06',
        'entite': 'ELCAM',
        'direction': 'Direction Generale',
        'departement': 'Operations',
        'fonction': 'Analyste',
        'categorie': 'Cadre moyen',
        'role': 'EMPLOYE',
        'n1_fonction': '',
        'ville': 'Douala',
        'contact_urgence': '+237655443322',
        'diplome': 'Master',
        'solde_conges': '7',
        'statut_matrimonial': 'Celibataire',
        'nombre_enfants': '1',
        'salaire_brut': '350000',
        'salaire_devise': 'XAF',
        'annee_experience': '4',
        'statut_employe': 'ACTIF',
        'type_contrat': 'CDD',
        'date_debut_contrat': '2026-02-06',
        'date_fin_contrat': '2026-12-31',
        'nouvelle_recrue': 'TRUE',
    }], mode='export')

    files = {
        'file': ('employees_styled.xlsx', stream.getvalue(), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }

    response = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200
    assert response.json()['imported_rows'] == 1

    check = client.get('/employees/8893', headers=auth_headers(9001, 'ADMIN'))
    assert check.status_code == 200
    payload = check.json()
    assert payload['contact_urgence'] == '+237655443322'
    assert payload['statut_matrimonial'] == 'Celibataire'
    assert payload['nombre_enfants'] == 1
    assert payload['salaire_devise'] == 'XAF'
    assert float(payload['salaire_brut']) == 350000.0
    assert payload['type_contrat'] == 'CDD'
    assert payload['date_debut_contrat'] == '2026-02-06'
    assert payload['date_fin_contrat'] == '2026-12-31'
    assert payload['nouvelle_recrue'] is True


def test_import_employees_multi_sheet_xlsx(client, seed_reference_data, auth_headers):
    """Test B : un xlsx avec plusieurs feuilles (Instructions/Référence + plusieurs feuilles d'employés)
    doit importer toutes les feuilles d'employés et ignorer les feuilles Instructions/Référence.
    Les employés déjà existants doivent être comptés dans skipped_rows, pas erreur.
    """
    rows_sheet1 = [
        {
            'matricule': 9101, 'nom': 'Multi', 'prenom': 'Un',
            'email': '9101@example.com', 'sexe': 'M',
            'date_embauche': date(2026, 2, 1),
            'entite': 'ELCAM', 'direction': 'Direction Generale',
            'departement': 'Operations', 'role': 'EMPLOYE', 'fonction': 'Analyste',
        },
        {
            'matricule': 9102, 'nom': 'Multi', 'prenom': 'Deux',
            'email': '9102@example.com', 'sexe': 'F',
            'date_embauche': date(2026, 2, 2),
            'entite': 'ELCAM', 'direction': 'Direction Generale',
            'departement': 'Operations', 'role': 'EMPLOYE', 'fonction': 'Analyste',
        },
    ]
    rows_sheet2 = [
        {
            'matricule': 9103, 'nom': 'Multi', 'prenom': 'Trois',
            'email': '9103@example.com', 'sexe': 'M',
            'date_embauche': date(2026, 2, 3),
            'entite': 'ELCAM', 'direction': 'Direction Generale',
            'departement': 'Operations', 'role': 'EMPLOYE', 'fonction': 'Analyste',
        },
    ]
    instructions = pd.DataFrame([{'col': 'Comment remplir le template'}, {'col': 'Ligne 2'}])
    reference = pd.DataFrame([{'code': 'M', 'libelle': 'Masculin'}])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(rows_sheet1).to_excel(writer, index=False, sheet_name='Employes-Site1')
        pd.DataFrame(rows_sheet2).to_excel(writer, index=False, sheet_name='Employes-Site2')
        instructions.to_excel(writer, index=False, sheet_name='Instructions')
        reference.to_excel(writer, index=False, sheet_name='Référence')
    output.seek(0)

    files = {
        'file': ('multi.xlsx', output.getvalue(),
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    r = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert r.status_code == 200, r.text
    payload = r.json()
    # 3 employés répartis sur 2 feuilles, Instructions/Référence ignorées
    assert payload['total_rows'] == 3, f"Attendu 3 lignes, obtenu {payload}"
    assert payload['imported_rows'] == 3
    assert payload['failed_rows'] == 0

    # Vérifier que les 3 ont bien été créés
    for m in (9101, 9102, 9103):
        chk = client.get(f'/employees/{m}', headers=auth_headers(9001, 'ADMIN'))
        assert chk.status_code == 200, f'Employé {m} introuvable'

    # Réimporter le même fichier → doit être idempotent (tous en skipped, aucun en failed)
    output.seek(0)
    files2 = {
        'file': ('multi.xlsx', output.getvalue(),
                 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    }
    r2 = client.post('/employees/import', files=files2, headers=auth_headers(9001, 'ADMIN'))
    assert r2.status_code == 200, r2.text
    p2 = r2.json()
    assert p2['imported_rows'] == 0, f"Réimport ne doit rien créer : {p2}"
    assert p2['skipped_rows'] == 3, f"Tous doivent être skipped : {p2}"
    assert p2['failed_rows'] == 0


def test_export_employees_csv(client, seed_reference_data, auth_headers):
    response = client.get('/employees/export?format=csv', headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200
    assert 'text/csv' in response.headers.get('content-type', '')
    assert 'attachment; filename=employees_export.csv' == response.headers.get('content-disposition')
    text = response.content.decode('utf-8')
    assert 'matricule,nom,prenom,email' in text
    assert 'contact_urgence' in text
    assert 'type_contrat' in text
    assert 'date_debut_contrat' in text
    assert 'date_fin_contrat' in text
    assert 'nouvelle_recrue' in text


def test_export_employees_xlsx(client, seed_reference_data, auth_headers):
    response = client.get('/employees/export?format=xlsx', headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200
    assert response.headers.get('content-disposition') == 'attachment; filename=employees_export.xlsx'
    assert len(response.content) > 100

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames[:3] == ['Employés', 'Instructions', 'Référence']

    ws = wb['Employés']
    assert str(ws['A1'].value).startswith('EXPORT EMPLOYÉS — EMS')
    expected_headers = [f"{label} *" if required else label for _key, label, required, _width, _note in COLUMNS]
    actual_headers = [ws.cell(row=3, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert actual_headers == expected_headers
    assert ws['A4'].value is not None


def test_import_employees_mdb(client, seed_reference_data, auth_headers, monkeypatch):
    from app.routers import employees as employees_router

    def fake_access_reader(content, filename, table_name=None):
        assert filename == 'employees.mdb'
        assert table_name is None
        return pd.DataFrame([
            {
                'matricule': 8890,
                'nom': 'Access',
                'prenom': 'User',
                'email': '8890@example.com',
                'sexe': 'M',
                'date_embauche': date(2026, 2, 3),
                'entite': 'ELCAM',
                'direction': 'Direction Generale',
                'departement': 'Operations',
                'role': 'EMPLOYE',
                'fonction': 'Analyste',
            }
        ]), 'Employees'

    monkeypatch.setattr(employees_router, 'read_access_dataframe', fake_access_reader)

    files = {
        'file': ('employees.mdb', b'fake-access-db', 'application/octet-stream')
    }

    response = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200
    assert response.json()['table'] == 'Employees'
    assert response.json()['imported_rows'] == 1


def test_import_employees_accdb_requires_table_selection(client, seed_reference_data, auth_headers, monkeypatch):
    from app.routers import employees as employees_router

    def fake_access_reader(content, filename, table_name=None):
        raise AccessImportError(
            code='access_table_required',
            message='Plusieurs tables Access detectees. Selectionnez une table.',
            available_tables=['Employes', 'Archives'],
            status_code=400,
        )

    monkeypatch.setattr(employees_router, 'read_access_dataframe', fake_access_reader)

    files = {
        'file': ('employees.accdb', b'fake-access-db', 'application/octet-stream')
    }

    response = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 400
    payload = response.json()['detail']
    assert payload['code'] == 'access_table_required'
    assert payload['available_tables'] == ['Employes', 'Archives']


def test_import_employees_accdb_with_selected_table(client, seed_reference_data, auth_headers, monkeypatch):
    from app.routers import employees as employees_router

    def fake_access_reader(content, filename, table_name=None):
        assert table_name == 'Employes'
        return pd.DataFrame([
            {
                'matricule': 8891,
                'nom': 'Access',
                'prenom': 'Selected',
                'email': '8891@example.com',
                'sexe': 'F',
                'date_embauche': date(2026, 2, 4),
                'entite': 'ELCAM',
                'direction': 'Direction Generale',
                'departement': 'Operations',
                'role': 'EMPLOYE',
                'fonction': 'Analyste',
            }
        ]), 'Employes'

    monkeypatch.setattr(employees_router, 'read_access_dataframe', fake_access_reader)

    files = {
        'file': ('employees.accdb', b'fake-access-db', 'application/octet-stream')
    }

    response = client.post('/employees/import?table=Employes', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 200
    assert response.json()['table'] == 'Employes'
    assert response.json()['imported_rows'] == 1


def test_import_employees_access_driver_error(client, seed_reference_data, auth_headers, monkeypatch):
    from app.routers import employees as employees_router

    def fake_access_reader(content, filename, table_name=None):
        raise AccessImportError(
            code='access_driver_unavailable',
            message='Aucun pilote ODBC Access compatible n est disponible sur ce serveur.',
            status_code=500,
        )

    monkeypatch.setattr(employees_router, 'read_access_dataframe', fake_access_reader)

    files = {
        'file': ('employees.accdb', b'fake-access-db', 'application/octet-stream')
    }

    response = client.post('/employees/import', files=files, headers=auth_headers(9001, 'ADMIN'))
    assert response.status_code == 500
    payload = response.json()['detail']
    assert payload['code'] == 'access_driver_unavailable'


def test_employee_localisation_and_geo_filters(client, seed_reference_data, auth_headers):
    payload = {
        'matricule': 8892,
        'nom': 'Geo',
        'prenom': 'Filter',
        'email': '8892@example.com',
        'sexe': 'M',
        'date_embauche': '2026-02-05',
        'entite': 'ELCAM',
        'direction': 'Direction Generale',
        'departement': 'Operations',
        'role': 'EMPLOYE',
        'fonction': 'Analyste',
        'ville': 'Douala',
        'contact_urgence': '+237677000000',
    }

    create = client.post('/employees/', json=payload)
    assert create.status_code == 200

    emp = create.json()
    assert emp['ville'] == 'Douala'
    assert emp['pays'] == 'Cameroun'
    assert emp['id_localisation'] is not None
    assert emp['id_pays'] == seed_reference_data['pays'].id_pays
    assert emp['contact_urgence'] == '+237677000000'

    by_country = client.get(f"/employees/?id_pays={seed_reference_data['pays'].id_pays}")
    assert by_country.status_code == 200
    assert any(int(row['matricule']) == 8892 for row in by_country.json())

    by_city = client.get(f"/employees/?id_localisation={seed_reference_data['localisation'].id_localisation}")
    assert by_city.status_code == 200
    assert any(int(row['matricule']) == 8892 for row in by_city.json())
