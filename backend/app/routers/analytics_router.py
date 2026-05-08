"""
Router â€” Export Excel multi-feuilles.

GET /api/analytics/export-excel
GÃ©nÃ¨re un fichier .xlsx avec 9 feuilles :
  1. EmployÃ©s              â€” liste complÃ¨te
  2. Absences_mois         â€” absences par mois
  3. Missions              â€” missions avec frais
  4. Scores_comportement   â€” scores par employÃ©
  5. Tendances_12mois      â€” tendances par type de demande sur 12 mois glissants
  6. Absenteisme_dept      â€” absences par dÃ©partement sur l'annÃ©e
  7. Solde_conges          â€” distribution par tranche de solde
  8. Formation             â€” taux de formation annuel
  9. Distribution_org      â€” rÃ©partition par entitÃ© / direction / dÃ©partement

GET /api/analytics/export-dashboard/{matricule}
GÃ©nÃ¨re un fichier .xlsx scopÃ© sur le matricule :
  1. Profil           â€” donnÃ©es personnelles
  2. Mes_operations   â€” opÃ©rations par type et statut
  3. Perimetre_KPIs   â€” KPIs pÃ©rimÃ¨tre managÃ©rial (si habilitÃ©)
  4. Org_distribution â€” distribution direction / dÃ©partement
"""
import io
from datetime import datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func

from .. import models
from ..db import get_db
from ..utils.security import get_current_user

router = APIRouter(prefix='/api/analytics', tags=['analytics'])

_ROLES_RH = {'RH', 'ADMIN', 'DIRECTEUR', 'DG', 'PCA'}


@router.get('/export-excel')
def export_excel(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Export Excel multi-feuilles â€” rÃ©servÃ© aux rÃ´les habilitants."""
    from fastapi import HTTPException
    role = (current_user.get('role') or '').upper()
    if role not in _ROLES_RH:
        raise HTTPException(status_code=403, detail='AccÃ¨s rÃ©servÃ©.')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl n'est pas installÃ© sur le serveur.")

    wb = openpyxl.Workbook()

    # â”€â”€ Couleurs charte â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    BRAND_BLUE = '02162e'
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(fill_type='solid', fgColor=BRAND_BLUE)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center')

    def _style_header(ws, headers: list):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # â”€â”€ Feuille 1 : EmployÃ©s â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws1 = wb.active
    ws1.title = 'EmployÃ©s'
    _style_header(ws1, ['Matricule', 'Nom', 'PrÃ©nom', 'Fonction', 'CatÃ©gorie',
                         'Statut', 'Date embauche', 'Direction', 'Email'])
    employes = db.query(models.Employe).all()
    dir_map = {d.id_direction: d.nom for d in db.query(models.Direction).all()}
    for e in employes:
        ws1.append([
            e.matricule,
            e.nom,
            e.prenom,
            e.fonction or '',
            e.categorie.value if e.categorie else '',
            e.statut_employe.value if e.statut_employe else '',
            e.date_embauche.isoformat() if e.date_embauche else '',
            dir_map.get(e.id_direction, ''),
            e.email or '',
        ])
    _autofit(ws1)

    # â”€â”€ Feuille 2 : Absences par mois â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2 = wb.create_sheet('Absences_mois')
    _style_header(ws2, ['AnnÃ©e', 'Mois', 'Type', 'Nb demandes', 'Nb validÃ©es', 'Nb rejetÃ©es'])
    rows_abs = (
        db.query(
            func.year(models.Operation.date_demande).label('annee'),
            func.month(models.Operation.date_demande).label('mois'),
            models.Operation.type_demande,
            func.count(models.Operation.id_operation).label('total'),
            func.sum(
                (models.Operation.statut == 'validÃ©').cast(models.Integer) if hasattr(models, 'Integer') else 0
            ).label('valides'),
        )
        .filter(models.Operation.type_demande.in_(['CongÃ©', 'Permission']))
        .group_by('annee', 'mois', models.Operation.type_demande)
        .order_by('annee', 'mois')
        .all()
    )
    # Calcul manuel pour Ã©viter les problÃ¨mes de cast
    absences_raw = (
        db.query(
            func.year(models.Operation.date_demande).label('annee'),
            func.month(models.Operation.date_demande).label('mois'),
            models.Operation.type_demande,
            models.Operation.statut,
            func.count(models.Operation.id_operation).label('cnt'),
        )
        .filter(models.Operation.type_demande.in_(['CongÃ©', 'Permission']))
        .group_by('annee', 'mois', models.Operation.type_demande, models.Operation.statut)
        .all()
    )
    # AgrÃ©gation Python
    abs_agg: dict = {}
    for row in absences_raw:
        key = (row.annee, row.mois, row.type_demande)
        if key not in abs_agg:
            abs_agg[key] = {'total': 0, 'valides': 0, 'rejetes': 0}
        abs_agg[key]['total'] += row.cnt
        if row.statut == 'validÃ©':
            abs_agg[key]['valides'] += row.cnt
        elif row.statut == 'rejetÃ©':
            abs_agg[key]['rejetes'] += row.cnt
    for (annee, mois, type_d), vals in sorted(abs_agg.items()):
        ws2.append([annee, mois, type_d, vals['total'], vals['valides'], vals['rejetes']])
    _autofit(ws2)

    # â”€â”€ Feuille 3 : Missions â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws3 = wb.create_sheet('Missions')
    _style_header(ws3, ['ID', 'Matricule', 'Nom', 'Pays', 'Ville', 'Date dÃ©part',
                         'Date retour', 'DurÃ©e (j)', 'Statut', 'Total frais'])
    missions_ops = (
        db.query(models.Operation, models.Mission)
        .join(models.Mission, models.Mission.id_mission == models.Operation.id_operation)
        .order_by(models.Operation.date_demande.desc())
        .all()
    )
    for op, m in missions_ops:
        # frais total
        fm = db.query(models.FraisMissionnaire).filter_by(id_mission=m.id_mission).all()
        total_frais = sum(float(f.total_frais or 0) for f in fm)
        emp_obj = db.query(models.Employe).filter_by(matricule=op.matricule).first()
        ws3.append([
            op.id_operation,
            op.matricule,
            f"{emp_obj.prenom} {emp_obj.nom}" if emp_obj else op.matricule,
            m.pays or '',
            m.ville or '',
            op.date_debut.isoformat() if op.date_debut else '',
            op.date_fin.isoformat() if op.date_fin else '',
            op.duree_jours or op.duree or '',
            op.statut,
            round(total_frais, 2),
        ])
    _autofit(ws3)

    # â”€â”€ Feuille 4 : Scores Comportementaux â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws4 = wb.create_sheet('Scores_comportement')
    _style_header(ws4, ['Matricule', 'Nom', 'PÃ©riode', 'Dimension', 'Score'])
    scores = (
        db.query(models.ScoreComportemental)
        .order_by(models.ScoreComportemental.periode.desc(), models.ScoreComportemental.matricule)
        .all()
    )
    emp_nom_map = {e.matricule: f"{e.prenom} {e.nom}" for e in db.query(models.Employe).all()}
    for s in scores:
        ws4.append([
            s.matricule,
            emp_nom_map.get(s.matricule, s.matricule),
            s.periode,
            s.dimension,
            float(s.valeur) if s.valeur is not None else '',
        ])
    _autofit(ws4)

    # â”€â”€ Feuille 5 : Tendances 12 mois glissants â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws5 = wb.create_sheet('Tendances_12mois')
    _style_header(ws5, ['AnnÃ©e', 'Mois', 'CongÃ©', 'Mission', 'Permission', 'Sortie', 'Total'])
    cutoff = datetime.utcnow() - timedelta(days=365)
    tendances_raw = (
        db.query(
            func.year(models.Operation.date_demande).label('annee'),
            func.month(models.Operation.date_demande).label('mois'),
            models.Operation.type_demande,
            func.count(models.Operation.id_operation).label('cnt'),
        )
        .filter(models.Operation.date_demande >= cutoff)
        .group_by('annee', 'mois', models.Operation.type_demande)
        .all()
    )
    tend_agg: dict = {}
    for r in tendances_raw:
        key = (r.annee, r.mois)
        if key not in tend_agg:
            tend_agg[key] = {'CongÃ©': 0, 'Mission': 0, 'Permission': 0, 'Sortie': 0}
        for t in ('CongÃ©', 'Mission', 'Permission', 'Sortie'):
            if r.type_demande == t:
                tend_agg[key][t] += r.cnt
    for (annee, mois), vals in sorted(tend_agg.items()):
        total = sum(vals.values())
        ws5.append([annee, mois, vals['CongÃ©'], vals['Mission'], vals['Permission'], vals['Sortie'], total])
    _autofit(ws5)

    # â”€â”€ Feuille 6 : AbsentÃ©isme par dÃ©partement â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws6 = wb.create_sheet('Absenteisme_dept')
    _style_header(ws6, ['DÃ©partement', 'Effectif actif', 'Jours absence (validÃ©s)', 'Taux (%)'])
    year_now = datetime.utcnow().year
    abs_dept_raw = (
        db.query(
            models.Departement.nom.label('dept_nom'),
            func.count(models.Operation.id_operation).label('nb_abs'),
        )
        .join(models.Employe, models.Operation.matricule == models.Employe.matricule)
        .join(models.Departement, models.Departement.dept_id == models.Employe.dept_id)
        .filter(
            models.Operation.type_demande.in_(['CongÃ©', 'Permission', 'Absence']),
            models.Operation.statut == 'approuvÃ©',
            func.year(models.Operation.date_demande) == year_now,
        )
        .group_by(models.Departement.nom)
        .all()
    )
    effectif_dept = dict(
        db.query(models.Departement.nom, func.count(models.Employe.matricule))
        .join(models.Employe, models.Employe.dept_id == models.Departement.dept_id)
        .filter(models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF)
        .group_by(models.Departement.nom)
        .all()
    )
    for dept_nom, nb_abs in sorted(abs_dept_raw, key=lambda x: x[1], reverse=True):
        eff = effectif_dept.get(dept_nom, 0) or 1
        taux = round(nb_abs / eff * 100, 1) if eff else 0
        ws6.append([dept_nom or 'N/A', effectif_dept.get(dept_nom, 0), nb_abs, taux])
    _autofit(ws6)

    # â”€â”€ Feuille 7 : Distribution solde congÃ©s â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws7 = wb.create_sheet('Solde_conges')
    _style_header(ws7, ['Tranche (jours)', 'Nombre d\'employÃ©s actifs'])
    tranches = [
        ('0 j', 0, 0),
        ('1 â€“ 5 j', 1, 5),
        ('6 â€“ 15 j', 6, 15),
        ('16 â€“ 25 j', 16, 25),
        ('26 â€“ 35 j', 26, 35),
        ('36 + j', 36, 9999),
    ]
    for label, lo, hi in tranches:
        q = db.query(func.count(models.Employe.matricule)).filter(
            models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF,
        )
        if lo == 0 and hi == 0:
            q = q.filter(models.Employe.solde_conges == 0)
        elif hi == 9999:
            q = q.filter(models.Employe.solde_conges > lo)
        else:
            q = q.filter(models.Employe.solde_conges >= lo, models.Employe.solde_conges <= hi)
        ws7.append([label, q.scalar() or 0])
    _autofit(ws7)

    # â”€â”€ Feuille 8 : Formation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws8 = wb.create_sheet('Formation')
    _style_header(ws8, ['AnnÃ©e', 'EmployÃ©s formÃ©s', 'Effectif actif', 'Taux (%)'])
    # Utilise les opÃ©rations de type Formation s'il en existe
    form_raw = (
        db.query(
            func.year(models.Operation.date_demande).label('annee'),
            func.count(func.distinct(models.Operation.matricule)).label('nb_formes'),
        )
        .filter(models.Operation.type_demande.ilike('%formation%'))
        .group_by('annee')
        .order_by('annee')
        .all()
    )
    actifs_total = db.query(func.count(models.Employe.matricule)).filter(
        models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF).scalar() or 1
    if form_raw:
        for r in form_raw:
            taux = round(r.nb_formes / actifs_total * 100, 1)
            ws8.append([r.annee, r.nb_formes, actifs_total, taux])
    else:
        # Fallback : chercher dans table Formation si elle existe
        try:
            form_table_raw = (
                db.query(
                    func.year(models.Formation.date_debut).label('annee'),
                    func.count(func.distinct(models.Inscription.matricule)).label('nb_formes'),
                )
                .join(models.Inscription, models.Inscription.id_formation == models.Formation.id_formation)
                .group_by('annee')
                .order_by('annee')
                .all()
            )
            for r in form_table_raw:
                taux = round(r.nb_formes / actifs_total * 100, 1)
                ws8.append([r.annee, r.nb_formes, actifs_total, taux])
        except Exception:
            ws8.append([year_now, 0, actifs_total, 0.0])
    _autofit(ws8)

    # â”€â”€ Feuille 9 : Distribution organisationnelle â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws9 = wb.create_sheet('Distribution_org')
    # Bloc entitÃ©s
    ws9.append(['=== Par entitÃ© ==='])
    ws9.append(['EntitÃ©', 'Effectif actif'])
    ent_rows = (
        db.query(models.Entite.nom, func.count(models.Employe.matricule))
        .join(models.Employe, models.Employe.id_entite == models.Entite.id_entite)
        .filter(models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF)
        .group_by(models.Entite.nom)
        .all()
    )
    for nom, cnt in ent_rows:
        ws9.append([nom, cnt])
    ws9.append([])
    # Bloc directions
    ws9.append(['=== Par direction ==='])
    ws9.append(['Direction', 'Effectif actif'])
    dir_rows = (
        db.query(models.Direction.nom, func.count(models.Employe.matricule))
        .join(models.Employe, models.Employe.id_direction == models.Direction.id_direction)
        .filter(models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF)
        .group_by(models.Direction.nom)
        .order_by(func.count(models.Employe.matricule).desc())
        .all()
    )
    for nom, cnt in dir_rows:
        ws9.append([nom, cnt])
    ws9.append([])
    # Bloc dÃ©partements
    ws9.append(['=== Par dÃ©partement ==='])
    ws9.append(['DÃ©partement', 'Effectif actif'])
    dept_rows = (
        db.query(models.Departement.nom, func.count(models.Employe.matricule))
        .join(models.Employe, models.Employe.dept_id == models.Departement.dept_id)
        .filter(models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF)
        .group_by(models.Departement.nom)
        .order_by(func.count(models.Employe.matricule).desc())
        .all()
    )
    for dept, cnt in dept_rows:
        ws9.append([dept or 'N/A', cnt])
    _autofit(ws9)

    # â”€â”€ SÃ©rialisation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"export_rh_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.get('/export-dashboard/{matricule}')
def export_dashboard(
    matricule: str,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Export Excel scopÃ© sur un matricule â€” accessible par l'employÃ© lui-mÃªme ou les RH."""
    role = (current_user.get('role') or '').upper()
    if current_user['matricule'] != matricule and role not in _ROLES_RH:
        raise HTTPException(status_code=403, detail='AccÃ¨s refusÃ©.')

    emp = db.query(models.Employe).filter_by(matricule=matricule).first()
    if not emp:
        raise HTTPException(status_code=404, detail='EmployÃ© introuvable.')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non disponible.")

    wb = openpyxl.Workbook()
    BRAND_BLUE = '02162e'
    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(fill_type='solid', fgColor=BRAND_BLUE)
    HEADER_ALIGN = Alignment(horizontal='center', vertical='center')

    def _hdr(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN

    def _autofit(ws):
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # â”€â”€ Feuille 1 : Profil â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws1 = wb.active
    ws1.title = 'Profil'
    dir_obj = db.query(models.Direction).filter_by(id_direction=emp.id_direction).first()
    ent_obj = db.query(models.Entite).filter_by(id_entite=emp.id_entite).first()
    dept_obj = db.query(models.Departement).filter_by(dept_id=emp.dept_id).first()
    rows_profil = [
        ('Matricule', emp.matricule),
        ('Nom', emp.nom),
        ('PrÃ©nom', emp.prenom),
        ('Fonction', emp.fonction or ''),
        ('CatÃ©gorie', emp.categorie.value if emp.categorie else ''),
        ('Statut', emp.statut_employe.value if emp.statut_employe else ''),
        ('Sexe', emp.sexe.value if emp.sexe else ''),
        ('Date de naissance', emp.date_naissance.isoformat() if emp.date_naissance else ''),
        ('Date embauche', emp.date_embauche.isoformat() if emp.date_embauche else ''),
        ('Direction', dir_obj.nom if dir_obj else ''),
        ('EntitÃ©', ent_obj.nom if ent_obj else ''),
        ('DÃ©partement', dept_obj.nom if dept_obj else ''),
        ('Email', emp.email or ''),
        ('Solde congÃ© (j)', emp.solde_conges or 0),
    ]
    for label, val in rows_profil:
        ws1.append([label, val])
    _autofit(ws1)

    # â”€â”€ Feuille 2 : Mes opÃ©rations â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws2 = wb.create_sheet('Mes_operations')
    _hdr(ws2, ['ID', 'Type', 'Date demande', 'Date dÃ©but', 'Date fin', 'DurÃ©e (j)', 'Statut'])
    ops = (
        db.query(models.Operation)
        .filter_by(matricule=matricule)
        .order_by(models.Operation.date_demande.desc())
        .all()
    )
    for op in ops:
        ws2.append([
            op.id_operation,
            op.type_demande,
            op.date_demande.isoformat() if op.date_demande else '',
            op.date_debut.isoformat() if op.date_debut else '',
            op.date_fin.isoformat() if op.date_fin else '',
            op.duree_jours or op.duree or '',
            op.statut,
        ])
    _autofit(ws2)

    # â”€â”€ Feuille 3 : PÃ©rimÃ¨tre KPIs (managers) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws3 = wb.create_sheet('Perimetre_KPIs')
    if role in _ROLES_RH:
        _hdr(ws3, ['Indicateur', 'Valeur'])
        total_emp = db.query(func.count(models.Employe.matricule)).scalar() or 0
        actifs = db.query(func.count(models.Employe.matricule)).filter(
            models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF).scalar() or 0
        hommes = db.query(func.count(models.Employe.matricule)).filter(
            models.Employe.sexe == models.SexeEnum.M).scalar() or 0
        femmes = db.query(func.count(models.Employe.matricule)).filter(
            models.Employe.sexe == models.SexeEnum.F).scalar() or 0
        conges_att = db.query(func.count(models.Operation.id_operation)).filter(
            models.Operation.type_demande == 'CongÃ©',
            models.Operation.statut == 'en attente').scalar() or 0
        missions_att = db.query(func.count(models.Operation.id_operation)).filter(
            models.Operation.type_demande == 'Mission',
            models.Operation.statut == 'en attente').scalar() or 0
        kpis = [
            ('Effectif total', total_emp),
            ('EmployÃ©s actifs', actifs),
            ('EmployÃ©s inactifs', total_emp - actifs),
            ('Hommes', hommes),
            ('Femmes', femmes),
            ('CongÃ©s en attente', conges_att),
            ('Missions en attente', missions_att),
        ]
        for label, val in kpis:
            ws3.append([label, val])
    else:
        ws3.append(['AccÃ¨s restreint', 'Non applicable'])
    _autofit(ws3)

    # â”€â”€ Feuille 4 : Distribution org (direction de l'employÃ©) â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    ws4 = wb.create_sheet('Org_distribution')
    _hdr(ws4, ['DÃ©partement', 'Effectif actif'])
    dept_rows = (
        db.query(models.Departement.nom, func.count(models.Employe.matricule))
        .join(models.Employe, models.Employe.dept_id == models.Departement.dept_id)
        .filter(
            models.Employe.id_direction == emp.id_direction,
            models.Employe.statut_employe == models.StatutEmployeEnum.ACTIF,
        )
        .group_by(models.Departement.nom)
        .all()
    )
    for dept, cnt in dept_rows:
        ws4.append([dept or 'N/A', cnt])
    _autofit(ws4)

    # â”€â”€ SÃ©rialisation â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nom_clean = f"{emp.nom}_{emp.prenom}".replace(' ', '_')
    filename = f"dashboard_{nom_clean}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.get('/recrues-par-an')
def recrues_par_an(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """Retourne le nombre de nouvelles recrues (nouvelle_recrue=True) par année d'embauche."""
    role = (current_user.get('role') or '').upper()
    if role not in _ROLES_RH:
        raise HTTPException(status_code=403, detail='Accès réservé.')

    rows = (
        db.query(
            func.year(models.Employe.date_embauche).label('annee'),
            func.count(models.Employe.matricule).label('count'),
        )
        .filter(
            models.Employe.nouvelle_recrue == True,  # noqa: E712
            models.Employe.date_embauche.isnot(None),
        )
        .group_by(func.year(models.Employe.date_embauche))
        .order_by(func.year(models.Employe.date_embauche))
        .all()
    )
    return [{'annee': r.annee, 'count': r.count} for r in rows if r.annee is not None]

