from __future__ import annotations

from io import BytesIO
import re
import unicodedata

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


BLEU = "02162E"
BLEU_CLAIR = "17283D"
ROUGE = "D0202B"
GRIS = "606060"
BLANC = "FFFFFF"
FOND_APP = "F4F5F7"
JAUNE_SOFT = "FFFDE7"
GRIS_EX = "EBEBEB"

FONT = "Century Gothic"

thin = Side(style="thin", color="D0D5DD")
medium = Side(style="medium", color=ROUGE)
b_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
b_req = Border(left=medium, right=medium, top=medium, bottom=medium)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


DB_ENTITES = ["ECG", "ELCAM", "EXCA"]

DB_DIRECTIONS = [
    "Audit Interne et Inspection Générale",
    "Conformité et Controle Interne",
    "Conseils et Financements Structurés",
    "Developpement et Investissement",
    "Direction de la Distribution",
    "Direction Financière et Comptable",
    "Organisation et Projets",
]

DB_DEPARTEMENTS = [
    "Affaires Juridiques & Fiscalité",
    "Audit interne",
    "Communication Marketing et Relations Publiques",
    "Comptabilité",
    "Controle de gestion",
    "Développement Commercial",
    "Développement commercial ELCAM",
    "Développement commercial EXCA",
    "Distribution Grandes Entreprises, Institutions et Fortunes",
    "Distribution particuliers et PME",
    "Financement & Structuration",
    "Gestion des Projets et Systèmes d'Informations",
    "Gestion et Analyse de portefeuille",
    "Inspection Generale",
    "Marketing Digital et Opérationnel",
    "Middle & Back Office",
    "Moyens Généraux",
    "Pool Grandes Entreprises & Fortunes",
    "Pool Particuliers & PME",
    "Ressources Humaines",
    "Trésorerie et Financement",
    "Trésorerie(ALM)",
]

DB_FONCTIONS = [
    "Administrateur Directeur Général",
    "Administrateur Général",
    "Analyste Financement et structuration",
    "Auditeur",
    "Chargé Administration Systèmes, Réseaux & Support IT",
    "Chargé Cloud et sécurité",
    "Chargé Communication",
    "Chargé Transformation Digitale, Innovation & Solutions Applicatives",
    "Chargé de négociation",
    "Chargé du développement portefeuille Grandes entreprise et Fortune",
    "Chargé du développement portefeuille particulier et PME",
    "Chargé développement Pool Grande Entreprise & Fortunes",
    "Chargé développement Pool Particuliers & PMEs",
    "DFC",
    "Directeur",
    "Directeur Audit Interne et Inspection Générale",
    "Directeur Conformité et Contrôle interne",
    "Directeur Conseil et Financement structurés",
    "Directeur Distribution",
    "Directeur Développement et investissement",
    "Directeur Général",
    "Directeur Général Adjoint",
    "Directeur des Organisations et projets",
    "Directeur financier et Comptable(DFC)",
    "Employé",
    "Inspecteur Générale(IG)",
    "PCA",
    "RH",
    "Représentants Résidents et responsables de la creation et relation d'affaires",
    "Responsable Comptable  Contrôle et Consolidation",
    "Responsable Des Resources Humaines",
    "Responsable Distribution Grandes Entreprises Institutions et Fortunes",
    "Responsable Distribution Particuliers et PME",
    "Responsable Département",
    "Responsable Financement et structuration",
    "Responsable Gestion et Analyste de portefeuille",
    "Responsable Middle & Back Office",
    "Responsable Trésorerie(ALM)",
    "Responsable affaires juridiques & fiscalité",
    "Responsable conformité et contrôle interne",
    "Responsable des Ressources Humaines",
    "Responsable des systèmes d'information",
    "Responsable du Développement",
    "Responsable développement Pool Grande Entreprise & Fortunes",
    "Responsable développement Pool Particuliers & PME",
    "Stagiaire académique",
    "Stagiaire professionnel",
    "chargé Analyste de portefeuille",
    "chargé Back Office & operations",
    "chargé community management accueil et courrier",
    "chargé de Gestions de portefeuille",
    "chargé de la fiscalité",
    "chargé des moyens généraux",
    "chargé des organisations et projets",
    "chargé des resources humaines",
    "chargé marketing digital opérationnel",
    "comptable",
    "comptable et responsable contrôle et consolidation",
    "contrôleur de gestion",
    "infographiste et déploiement",
    "responsable Trésorerie et financement",
    "responsable communication et relation publiques",
]

DB_ROLES = ["ADMIN", "DFC", "DG", "DIRECTEUR", "EMPLOYE", "PCA", "RESPONSABLE", "RH"]
DB_VILLES = ["Brazzaville", "Douala", "Libreville", "Yaoundé"]
DB_PAYS = ["Cameroun", "Congo", "Gabon", "Guinée équatoriale", "République centrafricaine", "Tchad"]

ENUMS = {
    "sexe": ["M", "F", "Autre"],
    "categorie": [
        "Cadre supérieur",
        "Cadre moyen",
        "Agent de maîtrise",
        "Agent qualifié",
        "Agent non qualifié",
        "Apprenti",
        "Stagiaire",
    ],
    "statut_matrimonial": ["Celibataire", "Marie"],
    "salaire_devise": ["XAF", "EUR", "USD", "XOF"],
    "statut_employe": ["ACTIF", "SUSPENDU", "CONGEDIE"],
    "type_contrat": ["CDI", "CDD", "Stagiaire"],
    "nouvelle_recrue": ["TRUE", "FALSE"],
}

COLUMNS = [
    ("matricule", "Matricule", True, 14, "Unique, alphanumérique (lettres, chiffres, tirets). Ex : EMP001"),
    ("nom", "Nom", True, 18, "Nom de famille. Ex : DUPONT"),
    ("prenom", "Prénom", True, 18, "Prénom. Ex : Jean"),
    ("email", "Email", False, 28, "Email professionnel unique. Ex : jean.dupont@elite.com"),
    ("telephone", "Téléphone", False, 16, "Format international : +237600000000"),
    ("sexe", "Sexe", False, 10, "M | F | Autre"),
    ("date_naissance", "Date de naissance", False, 16, "Format : aaaa-mm-jj  Ex : 1990-03-15"),
    ("date_embauche", "Date d'embauche", True, 16, "Format : aaaa-mm-jj  Ex : 2024-01-15  OBLIGATOIRE"),
    ("entite", "Entité", True, 14, "Valeur exacte de la BD : ECG | ELCAM | EXCA  OBLIGATOIRE"),
    ("direction", "Direction", False, 36, "Nom exact de la direction (voir liste déroulante)"),
    ("departement", "Département", False, 38, "Nom exact du département (voir liste déroulante)"),
    ("fonction", "Fonction / Poste", False, 36, "Intitulé de poste (voir liste déroulante ou saisir librement)"),
    ("categorie", "Catégorie", False, 22, "Cadre supérieur | Cadre moyen | Agent de maîtrise | Agent qualifié | Agent non qualifié | Apprenti | Stagiaire"),
    ("role", "Rôle applicatif", False, 16, "ADMIN | DFC | DG | DIRECTEUR | EMPLOYE | PCA | RESPONSABLE | RH"),
    ("n1_fonction", "Fonction du N+1", False, 36, "Fonction du manager direct — l'app retrouve son matricule automatiquement"),
    ("ville", "Ville", False, 14, "Ville de résidence / travail"),
    ("contact_urgence", "Contact d'urgence", False, 16, "Tél. de la personne à contacter en cas d'urgence"),
    ("diplome", "Diplôme", False, 14, "Ex : Master | Licence | BAC | BTS | BEPC"),
    ("solde_conges", "Solde congés (j)", False, 12, "Nombre de jours de congés initiaux. 0 par défaut"),
    ("statut_matrimonial", "Statut matrimonial", False, 18, "Celibataire | Marie"),
    ("nombre_enfants", "Nombre d'enfants", False, 12, "Nombre entier ≥ 0"),
    ("salaire_brut", "Salaire brut", False, 14, "Salaire brut mensuel (chiffre uniquement). Ex : 850000"),
    ("salaire_devise", "Devise", False, 10, "XAF | EUR | USD | XOF  (défaut : XAF)"),
    ("annee_experience", "Années d'exp.", False, 12, "Nombre entier"),
    ("statut_employe", "Statut employé", False, 14, "ACTIF | SUSPENDU | CONGEDIE  (défaut : ACTIF)"),
    ("type_contrat", "Type de contrat", False, 16, "CDI | CDD | Stagiaire"),
    ("date_debut_contrat", "Date de début contrat", False, 18, "Format : aaaa-mm-jj. Recommandé pour CDD et Stagiaire"),
    ("date_fin_contrat", "Date de fin contrat", False, 18, "Format : aaaa-mm-jj. Obligatoire pour CDD et Stagiaire"),
    ("nouvelle_recrue", "Nouvelle recrue", False, 14, "TRUE | FALSE. Comptabilisé dans les statistiques de recrutement"),
]

EXAMPLES = [
    {
        "matricule": "EMP001",
        "nom": "NGUEMA",
        "prenom": "Paul",
        "email": "p.nguema@elcam.com",
        "telephone": "+237690000001",
        "sexe": "M",
        "date_naissance": "1988-05-12",
        "date_embauche": "2020-03-01",
        "entite": "ELCAM",
        "direction": "Direction Financière et Comptable",
        "departement": "Comptabilité",
        "fonction": "Directeur financier et Comptable(DFC)",
        "categorie": "Cadre supérieur",
        "role": "DFC",
        "n1_fonction": "Administrateur Général",
        "ville": "Yaoundé",
        "contact_urgence": "",
        "diplome": "Master",
        "solde_conges": "18",
        "statut_matrimonial": "Marie",
        "nombre_enfants": "2",
        "salaire_brut": "1200000",
        "salaire_devise": "XAF",
        "annee_experience": "12",
        "statut_employe": "ACTIF",
        "type_contrat": "CDI",
        "date_debut_contrat": "2020-03-01",
        "date_fin_contrat": "",
        "nouvelle_recrue": "FALSE",
    },
    {
        "matricule": "STG002",
        "nom": "MBARGA",
        "prenom": "Chloé",
        "email": "c.mbarga@exca.com",
        "telephone": "+237699000002",
        "sexe": "F",
        "date_naissance": "2003-11-07",
        "date_embauche": "2026-04-01",
        "entite": "EXCA",
        "direction": "",
        "departement": "",
        "fonction": "Stagiaire académique",
        "categorie": "Stagiaire",
        "role": "EMPLOYE",
        "n1_fonction": "",
        "ville": "Douala",
        "contact_urgence": "",
        "diplome": "BAC",
        "solde_conges": "0",
        "statut_matrimonial": "Celibataire",
        "nombre_enfants": "0",
        "salaire_brut": "",
        "salaire_devise": "XAF",
        "annee_experience": "0",
        "statut_employe": "ACTIF",
        "type_contrat": "Stagiaire",
        "date_debut_contrat": "2026-04-01",
        "date_fin_contrat": "2026-09-30",
        "nouvelle_recrue": "TRUE",
    },
]

REF_COLS_DEF = [
    ("Entités", DB_ENTITES),
    ("Directions", DB_DIRECTIONS),
    ("Depts", DB_DEPARTEMENTS),
    ("Fonctions", DB_FONCTIONS),
    ("Rôles", DB_ROLES),
    ("Villes", DB_VILLES),
    ("Pays", DB_PAYS),
]

REF_MAP = {
    "entite": ("A", len(DB_ENTITES) + 2),
    "direction": ("B", len(DB_DIRECTIONS) + 2),
    "departement": ("C", len(DB_DEPARTEMENTS) + 2),
    "fonction": ("D", len(DB_FONCTIONS) + 2),
    "role": ("E", len(DB_ROLES) + 2),
    "ville": ("F", len(DB_VILLES) + 2),
    "n1_fonction": ("D", len(DB_FONCTIONS) + 2),
    "type_contrat": ("A", 1),
}


def column_keys():
    return [key for key, *_ in COLUMNS]


def _normalize_header(value: str) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = text.lower().strip().replace("*", " ")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


IMPORT_HEADER_ALIASES = {}
for key, label, *_ in COLUMNS:
    IMPORT_HEADER_ALIASES[_normalize_header(key)] = key
    IMPORT_HEADER_ALIASES[_normalize_header(label)] = key

IMPORT_HEADER_ALIASES.update({
    _normalize_header("n1"): "n1",
    _normalize_header("téléphone"): "telephone",
    _normalize_header("contact urgence"): "contact_urgence",
    _normalize_header("solde congés"): "solde_conges",
    _normalize_header("date embauche"): "date_embauche",
    _normalize_header("statut employe"): "statut_employe",
    _normalize_header("salaire devise"): "salaire_devise",
    _normalize_header("date début contrat"): "date_debut_contrat",
    _normalize_header("date fin contrat"): "date_fin_contrat",
})


def canonical_import_header(value: str) -> str | None:
    return IMPORT_HEADER_ALIASES.get(_normalize_header(value))


def make_employee_workbook(rows=None, mode="template", include_reference=True, include_instructions=True):
    rows = list(rows or [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employés"

    n_cols = len(COLUMNS)
    data_first = 4
    display_rows = rows if mode == "export" else EXAMPLES + rows
    if mode == "template":
        blank_count = max(300 - len(rows), 0)
    else:
        blank_count = 0 if display_rows else 1
    data_last = data_first + len(display_rows) + blank_count - 1

    title = (
        "TEMPLATE IMPORT EMPLOYÉS — EMS  |  À remplir puis importer via page Employés › Importer"
        if mode == "template"
        else "EXPORT EMPLOYÉS — EMS  |  Même structure visuelle que le template d'import"
    )
    subtitle = (
        "  Colonnes OBLIGATOIRES (en-tête rouge)  ·  Colonnes optionnelles (en-tête bleu)  "
        "·  Lignes grisées = exemples à effacer  ·  Dates : aaaa-mm-jj"
        if mode == "template"
        else "  Export réimportable  ·  Même ordre de colonnes que le template d'import  ·  Dates : aaaa-mm-jj"
    )

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1, value=title)
    t.fill = fill(BLEU)
    t.font = Font(name=FONT, size=13, bold=True, color=BLANC)
    t.alignment = aln("center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.fill = fill(BLEU_CLAIR)
    sub.font = Font(name=FONT, size=9, italic=True, color=BLANC)
    sub.alignment = aln("left")
    ws.row_dimensions[2].height = 18

    for ci, (key, label, required, width, _) in enumerate(COLUMNS, start=1):
        c = ws.cell(row=3, column=ci)
        c.value = f"{label} *" if required else label
        c.fill = fill(ROUGE if required else BLEU_CLAIR)
        c.font = Font(name=FONT, size=10, bold=True, color=BLANC)
        c.alignment = aln("center", wrap=True)
        c.border = b_req if required else b_thin
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 34
    ws.freeze_panes = "B4"

    for offset, row_data in enumerate(display_rows, start=0):
        row_no = data_first + offset
        is_example = mode == "template" and offset < len(EXAMPLES)
        for ci, (key, *_rest) in enumerate(COLUMNS, start=1):
            value = row_data.get(key, "") if isinstance(row_data, dict) else ""
            c = ws.cell(row=row_no, column=ci, value=value)
            c.fill = fill(GRIS_EX if is_example else (FOND_APP if row_no % 2 == 0 else BLANC))
            c.font = Font(name=FONT, size=9 if is_example else 10, italic=is_example, color=GRIS)
            c.border = b_thin
            c.alignment = aln("left")
        ws.row_dimensions[row_no].height = 20 if is_example else 18

    for ri in range(data_first + len(display_rows), data_last + 1):
        bg = FOND_APP if ri % 2 == 0 else BLANC
        for ci in range(1, n_cols + 1):
            c = ws.cell(row=ri, column=ci)
            c.fill = fill(bg)
            c.border = b_thin
            c.alignment = aln("left")
            c.font = Font(name=FONT, size=10, color=GRIS)
        ws.row_dimensions[ri].height = 18

    date_keys = {"date_naissance", "date_embauche", "date_debut_contrat", "date_fin_contrat"}
    date_cols = {ci for ci, (key, *_) in enumerate(COLUMNS, start=1) if key in date_keys}
    for ri in range(data_first, data_last + 1):
        for ci in date_cols:
            ws.cell(row=ri, column=ci).number_format = "YYYY-MM-DD"

    for key, values in ENUMS.items():
        ci = next((i for i, (col_key, *_) in enumerate(COLUMNS, start=1) if col_key == key), None)
        if not ci:
            continue
        ltr = get_column_letter(ci)
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(values) + '"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Valeur invalide",
            error=f"Valeurs : {' | '.join(values)}",
        )
        dv.sqref = f"{ltr}{data_first}:{ltr}{data_last}"
        ws.add_data_validation(dv)

    for key, (ref_ltr, ref_last) in REF_MAP.items():
        if key == "type_contrat":
            continue
        ci = next((i for i, (col_key, *_) in enumerate(COLUMNS, start=1) if col_key == key), None)
        if not ci:
            continue
        ltr = get_column_letter(ci)
        dv = DataValidation(
            type="list",
            formula1=f"=Référence!${ref_ltr}$2:${ref_ltr}${ref_last}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
        )
        dv.sqref = f"{ltr}{data_first}:{ltr}{data_last}"
        ws.add_data_validation(dv)

    ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}3"

    if include_instructions:
        _add_instructions_sheet(wb)
    if include_reference:
        _add_reference_sheet(wb)
    return wb


def workbook_bytes(rows=None, mode="template"):
    wb = make_employee_workbook(rows=rows, mode=mode)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _sec_title(sheet, row, text):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = sheet.cell(row=row, column=1, value=text)
    c.fill = fill(BLEU)
    c.font = Font(name=FONT, size=11, bold=True, color=BLANC)
    c.alignment = aln("left")
    sheet.row_dimensions[row].height = 24
    return row + 1


def _write_row(sheet, row, label, value="", bg=BLANC, bold_l=False):
    c1 = sheet.cell(row=row, column=1, value=label)
    c1.font = Font(name=FONT, size=10, bold=bold_l, color=BLEU if bold_l else GRIS)
    c1.fill = fill(bg)
    c1.alignment = aln("left", wrap=True)
    if value:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c2 = sheet.cell(row=row, column=2, value=value)
        c2.font = Font(name=FONT, size=10, color=GRIS)
        c2.fill = fill(bg)
        c2.alignment = aln("left", wrap=True)
    sheet.row_dimensions[row].height = 16
    return row + 1


def _add_instructions_sheet(wb):
    wi = wb.create_sheet(title="Instructions")
    wi.merge_cells("A1:F1")
    wi.cell(row=1, column=1, value="INSTRUCTIONS D'IMPORT DES EMPLOYÉS — EMS").fill = fill(BLEU)
    wi.cell(row=1, column=1).font = Font(name=FONT, size=14, bold=True, color=BLANC)
    wi.cell(row=1, column=1).alignment = aln("center")
    wi.row_dimensions[1].height = 38

    r = 3
    r = _sec_title(wi, r, "  1.  PROCÉDURE D'IMPORT")
    r = _write_row(wi, r, "Étape 1", "Effacez les lignes d'exemple grisées (lignes 4 et 5 de l'onglet Employés).")
    r = _write_row(wi, r, "Étape 2", "Remplissez l'onglet « Employés » avec vos données. Utilisez les listes déroulantes.")
    r = _write_row(wi, r, "Étape 3", "Enregistrez ce fichier en .xlsx.")
    r = _write_row(wi, r, "Étape 4", "Dans EMS : page Employés → bouton ⋯ → Importer → sélectionnez ce fichier.")
    r = _write_row(wi, r, "Étape 5", "Un résumé s'affiche : « N importé(s), M échec(s) ». Les erreurs indiquent la ligne concernée.")
    r = _write_row(wi, r, "Étape 6", "L'administrateur est notifié automatiquement pour chaque employé créé + un résumé global.")
    r = _write_row(wi, r, "Étape 7", "Créez ensuite les comptes utilisateurs : page Employés → sélectionner → Créer compte.")
    r += 1

    r = _sec_title(wi, r, "  2.  CHAMPS OBLIGATOIRES")
    for key, label, req, _w, note in COLUMNS:
        if req:
            r = _write_row(wi, r, f"  ► {label} *", note, bold_l=True)
    r += 1

    r = _sec_title(wi, r, "  3.  FORMAT DES DONNÉES")
    r = _write_row(wi, r, "Dates", "Format : aaaa-mm-jj  (ISO 8601)  Ex : 2024-01-15  — dd/mm/yyyy aussi accepté")
    r = _write_row(wi, r, "Téléphones", "Format international : +237600000000")
    r = _write_row(wi, r, "Matricule", "Unique, alphanumérique. Ex : EMP001  /  AG-042")
    r = _write_row(wi, r, "Email", "Unique dans l'application. Doublon → ligne refusée.")
    r = _write_row(wi, r, "Salaire brut", "Nombre sans espaces ni symboles. Ex : 850000  (pas « 850 000 XAF »)")
    r += 1

    r = _sec_title(wi, r, "  4.  VALEURS ACCEPTÉES — ÉNUMÉRATIONS FIXES")
    for ci, hdr in enumerate(["Colonne", "Valeurs autorisées"], start=1):
        c = wi.cell(row=r, column=ci, value=hdr)
        c.fill = fill(BLEU_CLAIR)
        c.font = Font(name=FONT, size=10, bold=True, color=BLANC)
        if ci == 2:
            wi.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        wi.row_dimensions[r].height = 18
    r += 1
    for i, (key, vals) in enumerate(ENUMS.items()):
        label = next(lbl for col_key, lbl, *_ in COLUMNS if col_key == key)
        r = _write_row(wi, r, label, " | ".join(vals), bg=FOND_APP if i % 2 == 0 else BLANC)
    r += 1

    r = _sec_title(wi, r, "  5.  RÈGLES MÉTIER")
    r = _write_row(wi, r, "Unicité", "Matricule et email doivent être uniques en base de données.")
    r = _write_row(wi, r, "Âge minimum", "≥ 18 ans sauf si catégorie = Stagiaire ou Apprenti.")
    r = _write_row(wi, r, "Entité", "Valeur EXACTE de la BD : ECG | ELCAM | EXCA")
    r = _write_row(wi, r, "Direction / Dpt", "Résolution par nom exact. Si non trouvé : champ ignoré (pas d'erreur).")
    r = _write_row(wi, r, "N+1 (Fonction N+1)", "L'app recherche un employé ACTIF avec cette fonction → remplit le supérieur auto.")
    r = _write_row(wi, r, "Contrats", "Pour CDD et Stagiaire, renseignez idéalement la date de début et obligatoirement la date de fin.")
    r = _write_row(wi, r, "Comptes", "L'import crée le dossier employé uniquement. Compte login créé séparément par l'Admin.")

    wi.column_dimensions["A"].width = 28
    for col in "BCDEF":
        wi.column_dimensions[col].width = 22


def _add_reference_sheet(wb):
    wr = wb.create_sheet(title="Référence")
    wr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REF_COLS_DEF))
    note_cell = wr.cell(
        row=1,
        column=1,
        value=(
            "  Valeurs réelles de la base de données EMS.  "
            "Utilisez les noms EXACTS pour Entité, Direction, Département, Fonction, Rôle.  "
            "Ces listes alimentent les listes déroulantes de l'onglet Employés."
        ),
    )
    note_cell.fill = fill(JAUNE_SOFT)
    note_cell.font = Font(name=FONT, size=9, italic=True, color="8A6D00")
    note_cell.alignment = aln("left", wrap=True)
    wr.row_dimensions[1].height = 30

    for ci, (hdr, _) in enumerate(REF_COLS_DEF, start=1):
        c = wr.cell(row=2, column=ci, value=hdr)
        c.fill = fill(BLEU)
        c.font = Font(name=FONT, size=10, bold=True, color=BLANC)
        c.alignment = aln("center")
        c.border = b_thin
        wr.column_dimensions[get_column_letter(ci)].width = 42
    wr.row_dimensions[2].height = 22

    for ci, (_, rows_data) in enumerate(REF_COLS_DEF, start=1):
        for ri, val in enumerate(rows_data, start=3):
            c = wr.cell(row=ri, column=ci, value=val)
            c.font = Font(name=FONT, size=10, color=GRIS)
            c.border = b_thin
            c.fill = fill(FOND_APP if ri % 2 == 0 else BLANC)
            c.alignment = aln("left")
