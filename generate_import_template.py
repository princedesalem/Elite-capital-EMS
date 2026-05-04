# -*- coding: utf-8 -*-
"""
Génère TEMPLATE_IMPORT_EMPLOYES_EMS.xlsx — fichier Excel standalone
à remplir hors application puis importer via le bouton « Importer »
de la page Employés (EMS Admin uniquement).

Toutes les valeurs de référence sont celles de la base de données EMS réelle.
Police : Century Gothic partout.

Exécuter :  python generate_import_template.py
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Charte EMS ───────────────────────────────────────────────────────────
BLEU       = "02162E"
BLEU_CLAIR = "17283D"
ROUGE      = "D0202B"
GRIS       = "606060"
BLANC      = "FFFFFF"
FOND_APP   = "F4F5F7"
JAUNE_SOFT = "FFFDE7"
GRIS_EX    = "EBEBEB"

FONT = "Century Gothic"

thin   = Side(style="thin",   color="D0D5DD")
medium = Side(style="medium", color=ROUGE)
b_thin = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
b_req  = Border(left=medium, right=medium, top=medium, bottom=medium)

def fill(hex_):
    return PatternFill("solid", fgColor=hex_)

def fnt(color=GRIS, size=10, bold=False):
    return Font(name=FONT, size=size, bold=bold, color=color)

def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ═══════════════════════════════════════════════════════════════════════════
# RÉFÉRENTIELS — valeurs réelles de la base de données EMS
# (Mis à jour depuis la BD au 29/04/2026 — relancer generate_import_template.py
#  après chaque modification structurelle de la BD pour maintenir à jour.)
# ═══════════════════════════════════════════════════════════════════════════

# Entités (table ENTITE)
DB_ENTITES = [
    "ECG",
    "ELCAM",
    "EXCA",
]

# Directions (table DIRECTION) — dédupliquées
DB_DIRECTIONS = [
    "Audit Interne et Inspection Générale",
    "Conformité et Controle Interne",
    "Conseils et Financements Structurés",
    "Developpement et Investissement",
    "Direction de la Distribution",
    "Direction Financière et Comptable",
    "Organisation et Projets",
]

# Départements (table DEPARTEMENT) — dédupliqués
DB_DEPARTEMENTS = [
    "Affaires Juridiques & Fiscalité",
    "Audit interne",
    "Communication Marketing et Relations Publiques",
    "Comptabilité",
    "Controle de gestion",
    "Développement Commercial",
    "Développement commercial ELCAM",
    "Développement commercial EXCA",
    "Distribution Grandes Entreprises, Institutions et Fortunes",
    "Distribution particuliers et PME",
    "Financement & Structuration",
    "Gestion des Projets et Systèmes d'Informations",
    "Gestion et Analyse de portefeuille",
    "Inspection Generale",
    "Marketing Digital et Opérationnel",
    "Middle & Back Office",
    "Moyens Généraux",
    "Pool Grandes Entreprises & Fortunes",
    "Pool Particuliers & PME",
    "Ressources Humaines",
    "Trésorerie et Financement",
    "Trésorerie(ALM)",
]

# Fonctions (table FONCTION_REFERENCE + EMPLOYE.fonction) — dédupliquées
DB_FONCTIONS = [
    "Administrateur Directeur Général",
    "Administrateur Général",
    "Analyste Financement et structuration",
    "Auditeur",
    "Chargé Administration Systèmes, Réseaux & Support IT",
    "Chargé Cloud et sécurité",
    "Chargé Communication",
    "Chargé Transformation Digitale, Innovation & Solutions Applicatives",
    "Chargé de négociation",
    "Chargé du développement portefeuille Grandes entreprise et Fortune",
    "Chargé du développement portefeuille particulier et PME",
    "Chargé développement Pool Grande Entreprise & Fortunes",
    "Chargé développement Pool Particuliers & PMEs",
    "DFC",
    "Directeur",
    "Directeur Audit Interne et Inspection Générale",
    "Directeur Conformité et Contrôle interne",
    "Directeur Conseil et Financement structurés",
    "Directeur Distribution",
    "Directeur Développement et investissement",
    "Directeur Général",
    "Directeur Général Adjoint",
    "Directeur des Organisations et projets",
    "Directeur financier et Comptable(DFC)",
    "Employé",
    "Inspecteur Générale(IG)",
    "PCA",
    "RH",
    "Représentants Résidents et responsables de la creation et relation d'affaires",
    "Responsable Comptable  Contrôle et Consolidation",
    "Responsable Des Resources Humaines",
    "Responsable Distribution Grandes Entreprises Institutions et Fortunes",
    "Responsable Distribution Particuliers et PME",
    "Responsable Département",
    "Responsable Financement et structuration",
    "Responsable Gestion et Analyste de portefeuille",
    "Responsable Middle & Back Office",
    "Responsable Trésorerie(ALM)",
    "Responsable affaires juridiques & fiscalité",
    "Responsable conformité et contrôle interne",
    "Responsable des Ressources Humaines",
    "Responsable des systèmes d'information",
    "Responsable du Développement",
    "Responsable développement Pool Grande Entreprise & Fortunes",
    "Responsable développement Pool Particuliers & PME",
    "Stagiaire académique",
    "Stagiaire professionnel",
    "chargé Analyste de portefeuille",
    "chargé Back Office & operations",
    "chargé community management accueil et courrier",
    "chargé de Gestions de portefeuille",
    "chargé de la fiscalité",
    "chargé des moyens généraux",
    "chargé des organisations et projets",
    "chargé des resources humaines",
    "chargé marketing digital opérationnel",
    "comptable",
    "comptable et responsable contrôle et consolidation",
    "contrôleur de gestion",
    "infographiste et déploiement",
    "responsable Trésorerie et financement",
    "responsable communication et relation publiques",
]

# Rôles applicatifs (table roles)
DB_ROLES = [
    "ADMIN",
    "DFC",
    "DG",
    "DIRECTEUR",
    "EMPLOYE",
    "PCA",
    "RESPONSABLE",
    "RH",
]

# Villes (table LOCALISATION)
DB_VILLES = [
    "Brazzaville",
    "Douala",
    "Libreville",
    "Yaoundé",
]

# Pays (table PAYS)
DB_PAYS = [
    "Cameroun",
    "Congo",
    "Gabon",
    "Guinée équatoriale",
    "République centrafricaine",
    "Tchad",
]

# ── Énumérations fixes (enum SQLAlchemy dans models.py) ──────────────────
ENUMS = {
    "sexe":               ["M", "F", "Autre"],
    "categorie":          ["Cadre supérieur", "Cadre moyen", "Agent de maîtrise",
                           "Agent qualifié", "Agent non qualifié", "Apprenti", "Stagiaire"],
    "statut_matrimonial": ["Celibataire", "Marie"],
    "salaire_devise":     ["XAF", "EUR", "USD", "XOF"],
    "statut_employe":     ["ACTIF", "SUSPENDU", "CONGEDIE"],
}

# ── Définition des colonnes (ordre = formulaire EmployeeForm.jsx) ─────────
COLUMNS = [
    # (clé_api, label_excel, obligatoire, largeur_col, note_aide)
    ("matricule",         "Matricule",            True,  14, "Unique, alphanumérique (lettres, chiffres, tirets). Ex : EMP001"),
    ("nom",               "Nom",                  True,  18, "Nom de famille. Ex : DUPONT"),
    ("prenom",            "Prénom",               True,  18, "Prénom. Ex : Jean"),
    ("email",             "Email",                False, 28, "Email professionnel unique. Ex : jean.dupont@elite.com"),
    ("telephone",         "Téléphone",            False, 16, "Format international : +237600000000"),
    ("sexe",              "Sexe",                 False, 10, "M | F | Autre"),
    ("date_naissance",    "Date de naissance",    False, 16, "Format : aaaa-mm-jj  Ex : 1990-03-15"),
    ("date_embauche",     "Date d'embauche",      True,  16, "Format : aaaa-mm-jj  Ex : 2024-01-15  OBLIGATOIRE"),
    ("entite",            "Entité",               True,  14, "Valeur exacte de la BD : ECG | ELCAM | EXCA  OBLIGATOIRE"),
    ("direction",         "Direction",            False, 36, "Nom exact de la direction (voir liste déroulante)"),
    ("departement",       "Département",          False, 38, "Nom exact du département (voir liste déroulante)"),
    ("fonction",          "Fonction / Poste",     False, 36, "Intitulé de poste (voir liste déroulante ou saisir librement)"),
    ("categorie",         "Catégorie",            False, 22, "Cadre supérieur | Cadre moyen | Agent de maîtrise | Agent qualifié | Agent non qualifié | Apprenti | Stagiaire"),
    ("role",              "Rôle applicatif",      False, 16, "ADMIN | DFC | DG | DIRECTEUR | EMPLOYE | PCA | RESPONSABLE | RH"),
    ("n1_fonction",       "Fonction du N+1",      False, 36, "Fonction du manager direct — l'app retrouve son matricule automatiquement"),
    ("ville",             "Ville",                False, 14, "Ville de résidence / travail"),
    ("contact_urgence",   "Contact d'urgence",    False, 16, "Tél. de la personne à contacter en cas d'urgence"),
    ("diplome",           "Diplôme",              False, 14, "Ex : Master | Licence | BAC | BTS | BEPC"),
    ("solde_conges",      "Solde congés (j)",     False, 12, "Nombre de jours de congés initiaux. 0 par défaut"),
    ("statut_matrimonial","Statut matrimonial",   False, 18, "Celibataire | Marie"),
    ("nombre_enfants",    "Nombre d'enfants",     False, 12, "Nombre entier ≥ 0"),
    ("salaire_brut",      "Salaire brut",         False, 14, "Salaire brut mensuel (chiffre uniquement). Ex : 850000"),
    ("salaire_devise",    "Devise",               False, 10, "XAF | EUR | USD | XOF  (défaut : XAF)"),
    ("annee_experience",  "Années d'exp.",        False, 12, "Nombre entier"),
    ("statut_employe",    "Statut employé",       False, 14, "ACTIF | SUSPENDU | CONGEDIE  (défaut : ACTIF)"),
]

N_COLS     = len(COLUMNS)
DATA_FIRST = 4      # Ligne 1 = titre, 2 = sous-titre, 3 = en-têtes
N_DATA     = 300    # Lignes pré-équipées
DATA_LAST  = DATA_FIRST + N_DATA - 1

# Mapping clé → index de colonne (1-based)
col_idx = {key: ci for ci, (key, *_) in enumerate(COLUMNS, start=1)}

# ── Exemples (2 lignes grisées à effacer avant import) ────────────────────
EXAMPLES = [
    ("EMP001",  "NGUEMA",   "Paul",   "p.nguema@elcam.com",  "+237690000001",
     "M",  "1988-05-12", "2020-03-01", "ELCAM",
     "Direction Financière et Comptable", "Comptabilité",
     "Directeur financier et Comptable(DFC)", "Cadre supérieur", "DFC",
     "Administrateur Général", "Yaoundé", "", "Master",
     "18", "Marie", "2", "1200000", "XAF", "12", "ACTIF"),
    ("STG002",  "MBARGA",   "Chloé",  "c.mbarga@exca.com",   "+237699000002",
     "F",  "2003-11-07", "2026-04-01", "EXCA",
     "", "", "Stagiaire académique", "Stagiaire", "EMPLOYE",
     "", "Douala", "", "BAC",
     "0", "Celibataire", "0", "", "XAF", "0", "ACTIF"),
]

# ═══════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# ONGLET 1 — Employés
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Employés"

# Ligne 1 — Titre
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=N_COLS)
t = ws.cell(row=1, column=1,
    value="TEMPLATE IMPORT EMPLOYÉS — EMS  |  À remplir puis importer via page Employés › Importer")
t.fill = fill(BLEU)
t.font = Font(name=FONT, size=13, bold=True, color=BLANC)
t.alignment = aln("center")
ws.row_dimensions[1].height = 34

# Ligne 2 — Légende
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=N_COLS)
sub = ws.cell(row=2, column=1,
    value="  Colonnes OBLIGATOIRES (en-tête rouge)  ·  Colonnes optionnelles (en-tête bleu)  "
          "·  Lignes grisées = exemples à effacer  ·  Dates : aaaa-mm-jj")
sub.fill = fill(BLEU_CLAIR)
sub.font = Font(name=FONT, size=9, italic=True, color=BLANC)
sub.alignment = aln("left")
ws.row_dimensions[2].height = 18

# Ligne 3 — En-têtes
for ci, (key, label, required, width, _) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=3, column=ci)
    c.value     = f"{label} *" if required else label
    c.fill      = fill(ROUGE if required else BLEU_CLAIR)
    c.font      = Font(name=FONT, size=10, bold=True, color=BLANC)
    c.alignment = aln("center", wrap=True)
    c.border    = b_req if required else b_thin
    ws.column_dimensions[get_column_letter(ci)].width = width
ws.row_dimensions[3].height = 34
ws.freeze_panes = "B4"

# Lignes d'exemple (grisées)
for ei, ex in enumerate(EXAMPLES):
    row = DATA_FIRST + ei
    for ci, val in enumerate(ex, start=1):
        c = ws.cell(row=row, column=ci, value=val)
        c.fill      = fill(GRIS_EX)
        c.font      = Font(name=FONT, size=9, italic=True, color=GRIS)
        c.border    = b_thin
        c.alignment = aln("left")
    ws.row_dimensions[row].height = 20

# Lignes de données vides
for ri in range(DATA_FIRST + len(EXAMPLES), DATA_LAST + 1):
    bg = FOND_APP if ri % 2 == 0 else BLANC
    for ci in range(1, N_COLS + 1):
        c = ws.cell(row=ri, column=ci)
        c.fill   = fill(bg)
        c.border = b_thin
        c.alignment = aln("left")
        c.font   = Font(name=FONT, size=10, color=GRIS)
    ws.row_dimensions[ri].height = 18

# Format date sur colonnes date
date_cols = {ci for ci, (key, *_) in enumerate(COLUMNS, start=1)
             if key in ("date_naissance", "date_embauche")}
for ri in range(DATA_FIRST, DATA_LAST + 1):
    for ci in date_cols:
        ws.cell(row=ri, column=ci).number_format = "YYYY-MM-DD"

# ── Listes déroulantes — énumérations statiques ───────────────────────────
for key, values in ENUMS.items():
    ci = col_idx.get(key)
    if not ci:
        continue
    ltr = get_column_letter(ci)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(values) + '"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Valeur invalide",
        error=f"Valeurs : {' | '.join(values)}",
    )
    dv.sqref = f"{ltr}{DATA_FIRST}:{ltr}{DATA_LAST}"
    ws.add_data_validation(dv)

# ── Listes déroulantes — référentiels BD (via onglet Référence) ───────────
# Colonnes référentielles pointant vers l'onglet Référence
# col Référence : A=Entités  B=Directions  C=Depts  D=Fonctions  E=Rôles  F=Villes  G=Pays
REF_MAP = {
    "entite":      ("A", len(DB_ENTITES)    + 2),
    "direction":   ("B", len(DB_DIRECTIONS) + 2),
    "departement": ("C", len(DB_DEPARTEMENTS) + 2),
    "fonction":    ("D", len(DB_FONCTIONS)  + 2),
    "role":        ("E", len(DB_ROLES)      + 2),
    "ville":       ("F", len(DB_VILLES)     + 2),
    "n1_fonction": ("D", len(DB_FONCTIONS)  + 2),  # mêmes valeurs que Fonction
}
for key, (ref_ltr, ref_last) in REF_MAP.items():
    ci = col_idx.get(key)
    if not ci:
        continue
    ltr = get_column_letter(ci)
    dv = DataValidation(
        type="list",
        formula1=f"=Référence!${ref_ltr}$2:${ref_ltr}${ref_last}",
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=False,
    )
    dv.sqref = f"{ltr}{DATA_FIRST}:{ltr}{DATA_LAST}"
    ws.add_data_validation(dv)

ws.auto_filter.ref = f"A3:{get_column_letter(N_COLS)}3"

# ═══════════════════════════════════════════════════════════════════════════
# ONGLET 2 — Instructions
# ═══════════════════════════════════════════════════════════════════════════
wi = wb.create_sheet(title="Instructions")

def sec_title(sheet, row, text):
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    c = sheet.cell(row=row, column=1, value=text)
    c.fill = fill(BLEU)
    c.font = Font(name=FONT, size=11, bold=True, color=BLANC)
    c.alignment = aln("left")
    sheet.row_dimensions[row].height = 24
    return row + 1

def write_row(sheet, row, label, value="", bg=BLANC, bold_l=False):
    c1 = sheet.cell(row=row, column=1, value=label)
    c1.font      = Font(name=FONT, size=10, bold=bold_l, color=BLEU if bold_l else GRIS)
    c1.fill      = fill(bg)
    c1.alignment = aln("left", wrap=True)
    if value:
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        c2 = sheet.cell(row=row, column=2, value=value)
        c2.font      = Font(name=FONT, size=10, color=GRIS)
        c2.fill      = fill(bg)
        c2.alignment = aln("left", wrap=True)
    sheet.row_dimensions[row].height = 16
    return row + 1

# Titre
wi.merge_cells("A1:F1")
wi.cell(row=1, column=1, value="INSTRUCTIONS D'IMPORT DES EMPLOYÉS — EMS").fill = fill(BLEU)
wi.cell(row=1, column=1).font = Font(name=FONT, size=14, bold=True, color=BLANC)
wi.cell(row=1, column=1).alignment = aln("center")
wi.row_dimensions[1].height = 38

r = 3
r = sec_title(wi, r, "  1.  PROCÉDURE D'IMPORT")
r = write_row(wi, r, "Étape 1", "Effacez les lignes d'exemple grisées (lignes 4 et 5 de l'onglet Employés).")
r = write_row(wi, r, "Étape 2", "Remplissez l'onglet « Employés » avec vos données. Utilisez les listes déroulantes.")
r = write_row(wi, r, "Étape 3", "Enregistrez ce fichier en .xlsx.")
r = write_row(wi, r, "Étape 4", "Dans EMS : page Employés → bouton ⋯ → Importer → sélectionnez ce fichier.")
r = write_row(wi, r, "Étape 5", "Un résumé s'affiche : « N importé(s), M échec(s) ». Les erreurs indiquent la ligne concernée.")
r = write_row(wi, r, "Étape 6", "L'administrateur est notifié automatiquement pour chaque employé créé + un résumé global.")
r = write_row(wi, r, "Étape 7", "Créez ensuite les comptes utilisateurs : page Employés → sélectionner → Créer compte.")
r += 1

r = sec_title(wi, r, "  2.  CHAMPS OBLIGATOIRES")
for (key, label, req, w, note) in COLUMNS:
    if req:
        r = write_row(wi, r, f"  ► {label} *", note, bold_l=True)
r += 1

r = sec_title(wi, r, "  3.  FORMAT DES DONNÉES")
r = write_row(wi, r, "Dates",       "Format : aaaa-mm-jj  (ISO 8601)  Ex : 2024-01-15  — dd/mm/yyyy aussi accepté")
r = write_row(wi, r, "Téléphones",  "Format international : +237600000000")
r = write_row(wi, r, "Matricule",   "Unique, alphanumérique. Ex : EMP001  /  AG-042")
r = write_row(wi, r, "Email",       "Unique dans l'application. Doublon → ligne refusée.")
r = write_row(wi, r, "Salaire brut","Nombre sans espaces ni symboles. Ex : 850000  (pas « 850 000 XAF »)")
r += 1

r = sec_title(wi, r, "  4.  VALEURS ACCEPTÉES — ÉNUMÉRATIONS FIXES")
# Header tableau
for ci, hdr in enumerate(["Colonne", "Valeurs autorisées"], start=1):
    c = wi.cell(row=r, column=ci, value=hdr)
    c.fill = fill(BLEU_CLAIR)
    c.font = Font(name=FONT, size=10, bold=True, color=BLANC)
    if ci == 2:
        wi.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    wi.row_dimensions[r].height = 18
r += 1
for i, (key, vals) in enumerate(ENUMS.items()):
    label = next(lbl for (k, lbl, *_) in COLUMNS if k == key)
    r = write_row(wi, r, label, " | ".join(vals), bg=FOND_APP if i % 2 == 0 else BLANC)
r += 1

r = sec_title(wi, r, "  5.  RÈGLES MÉTIER")
r = write_row(wi, r, "Unicité",            "Matricule et email doivent être uniques en base de données.")
r = write_row(wi, r, "Âge minimum",        "≥ 18 ans sauf si catégorie = Stagiaire ou Apprenti.")
r = write_row(wi, r, "Entité",             "Valeur EXACTE de la BD : ECG | ELCAM | EXCA")
r = write_row(wi, r, "Direction / Dpt",    "Résolution par nom exact. Si non trouvé : champ ignoré (pas d'erreur).")
r = write_row(wi, r, "N+1 (Fonction N+1)", "L'app recherche un employé ACTIF avec cette fonction → remplit le supérieur auto.")
r = write_row(wi, r, "Comptes",            "L'import crée le dossier employé uniquement. Compte login créé séparément par l'Admin.")

wi.column_dimensions["A"].width = 28
for col in "BCDEF":
    wi.column_dimensions[col].width = 22

# ═══════════════════════════════════════════════════════════════════════════
# ONGLET 3 — Référence (données BD réelles)
# ═══════════════════════════════════════════════════════════════════════════
wr = wb.create_sheet(title="Référence")

REF_COLS_DEF = [
    ("Entités",    DB_ENTITES),
    ("Directions", DB_DIRECTIONS),
    ("Depts",      DB_DEPARTEMENTS),
    ("Fonctions",  DB_FONCTIONS),
    ("Rôles",      DB_ROLES),
    ("Villes",     DB_VILLES),
    ("Pays",       DB_PAYS),
]

# Ligne 1 — note
wr.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(REF_COLS_DEF))
note_cell = wr.cell(row=1, column=1,
    value="  Valeurs réelles de la base de données EMS (29/04/2026).  "
          "Utilisez les noms EXACTS pour Entité, Direction, Département, Fonction, Rôle.  "
          "Ces listes alimentent les listes déroulantes de l'onglet Employés.")
note_cell.fill = fill(JAUNE_SOFT)
note_cell.font = Font(name=FONT, size=9, italic=True, color="8A6D00")
note_cell.alignment = aln("left", wrap=True)
wr.row_dimensions[1].height = 30

# En-têtes colonnes référence
for ci, (hdr, _) in enumerate(REF_COLS_DEF, start=1):
    c = wr.cell(row=2, column=ci, value=hdr)
    c.fill = fill(BLEU)
    c.font = Font(name=FONT, size=10, bold=True, color=BLANC)
    c.alignment = aln("center")
    c.border    = b_thin
    wr.column_dimensions[get_column_letter(ci)].width = 42
wr.row_dimensions[2].height = 22

# Données
for ci, (_, rows_data) in enumerate(REF_COLS_DEF, start=1):
    for ri, val in enumerate(rows_data, start=3):
        c = wr.cell(row=ri, column=ci, value=val)
        c.font      = Font(name=FONT, size=10, color=GRIS)
        c.border    = b_thin
        c.fill      = fill(FOND_APP if ri % 2 == 0 else BLANC)
        c.alignment = aln("left")

# ── Sauvegarde ────────────────────────────────────────────────────────────
out_path = (
    r"c:\Users\cedri\OneDrive - ELITE CAPITAL Group S.A"
    r"\Documents\EMS\extranet\TEMPLATE_IMPORT_EMPLOYES_EMS.xlsx"
)
try:
    wb.save(out_path)
    print(f"Saved  → {out_path}")
except PermissionError:
    out_path = out_path.replace(".xlsx", "_v2.xlsx")
    wb.save(out_path)
    print(f"Fichier occupé — enregistré sous : {out_path}")

print(f"  Entités      : {len(DB_ENTITES)} valeurs")
print(f"  Directions   : {len(DB_DIRECTIONS)} valeurs")
print(f"  Départements : {len(DB_DEPARTEMENTS)} valeurs")
print(f"  Fonctions    : {len(DB_FONCTIONS)} valeurs")
print(f"  Rôles        : {len(DB_ROLES)} valeurs")
print(f"  Villes       : {len(DB_VILLES)} valeurs")
print(f"  Pays         : {len(DB_PAYS)} valeurs")
print(f"  Police       : {FONT}")
print(f"  Lignes       : {N_DATA} lignes pré-équipées  ·  {len(EXAMPLES)} lignes d'exemple grisées")
